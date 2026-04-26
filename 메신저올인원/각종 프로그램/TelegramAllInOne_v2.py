#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════╗
║       텔레그램 올인원 자동화  v2.0  (Single-File Edition)       ║
║  Telegram All-in-One Automation v2.0                         ║
║  ─────────────────────────────────────────────────────────  ║
║  구조:                                                       ║
║    [A] Foundation  — 상수/유틸/StopController                ║
║    [B] Data Layer  — 링크파서/ExcelManager/이미지유틸          ║
║    [C] Workers     — JoinWorker / MessageWorker              ║
║    [D] UI Layer    — Widgets / CoordTab / JoinOnlyTab /      ║
║                      JoinTab / MessageTab / ManagerTab       ║
║    [E] Entry       — TelegramAllInOne / main()               ║
╚══════════════════════════════════════════════════════════════╝
"""

import sys
import os
import re
import time
import random
import threading
import json
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, simpledialog
from typing import List, Optional, Dict
from datetime import datetime
from dataclasses import dataclass

import pyautogui
import pyperclip
import keyboard
import openpyxl

# 프로젝트 루트를 sys.path 에 추가 (PyInstaller 환경 대응)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))



# ════════════════════════════════════════════════════════════
# [A] Foundation Layer — 상수 / 유틸 / StopController
# ════════════════════════════════════════════════════════════

# [DUP_REMOVED] import os
# [DUP_REMOVED] import sys

# ─── 테마 색상 ───────────────────────────────────────────
THEME = {
    "bg":      "#1a1a2e",
    "card":    "#16213e",
    "border":  "#0f3460",
    "accent":  "#2b8be8",
    "text":    "#eaeaea",
    "subtext": "#a0a0c0",
    "success": "#4ecca3",
    "warning": "#f5a623",
    "danger":  "#e94560",
    "input_bg":"#0d1b2a",
}

# ─── 폰트 ────────────────────────────────────────────────
FONT = {
    "title":   ("맑은 고딕", 13, "bold"),
    "label":   ("맑은 고딕", 10),
    "small":   ("맑은 고딕", 9),
    "mono":    ("Consolas", 9),
    "btn":     ("맑은 고딕", 10, "bold"),
    "header":  ("맑은 고딕", 11, "bold"),
}

# ─── 기본 딜레이 (초) ────────────────────────────────────
DEFAULT_DELAYS = {
    "chrome_load":               2.0,
    "telegram_open":             1.5,
    "join_click":                2.0,
    "join_result_check_delay":   2.5,   # JOIN 클릭 후 픽셀 판정 전 대기
    "join_result_check_retry":   3,     # 픽셀 판정 재시도 횟수
    "after_type":                0.5,
    "after_send":                1.0,
    "after_back":                0.8,
    "between_min":               3.0,
    "between_max":               7.0,
    "file_dialog_open":          1.5,
    "after_addr_click":          0.5,
    "after_path_input":          0.8,
    "after_file_enter":          2.0,
}

# ─── 상태값 ──────────────────────────────────────────────
STATUS = {
    "UNJOINED":   "미가입",
    "JOINED":     "가입완료",
    "SENT":       "발송완료",
    "FAILED":     "실패",
    "SKIPPED":    "건너뜀",
}

STATUS_COLORS = {
    "미가입":  "#a0a0c0",
    "가입완료": "#f5a623",
    "발송완료": "#4ecca3",
    "실패":    "#e94560",
    "건너뜀":  "#888888",
}

# ─── Excel 컬럼 인덱스 (1-based) ─────────────────────────
EXCEL_COL = {
    "link":       1,
    "status":     2,
    "custom_msg": 3,
    "image_path": 4,
    "processed":  5,
    "note":       6,
}

EXCEL_HEADERS = ["링크", "상태", "커스텀메시지", "이미지경로", "처리일시", "비고"]

# ─── 경로 유틸 ───────────────────────────────────────────
def get_base_path() -> str:
    """실행 파일 기준 경로 (PyInstaller 대응)"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_presets_dir(tab_name: str = "coord") -> str:
    """프리셋 폴더 경로 (없으면 생성)"""
    path = os.path.join(get_base_path(), "presets", tab_name)
    os.makedirs(path, exist_ok=True)
    return path


def get_config_path(filename: str = "config_telegram.json") -> str:
    """설정 파일 전체 경로"""
    return os.path.join(get_base_path(), filename)


def get_outputs_dir() -> str:
    """출력 폴더 경로"""
    path = os.path.join(get_base_path(), "outputs")
    os.makedirs(path, exist_ok=True)
    return path


# ── A2: helpers.py ──────────────────────────────────
# [DUP_REMOVED] import time
# [DUP_REMOVED] import random
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import pyperclip
# [DUP_REMOVED] import pyautogui
# [DUP_REMOVED] import keyboard
# [DUP_REMOVED] from datetime import datetime


# ─── 타입 변환 ───────────────────────────────────────────
def safe_int(val, default: int = 0) -> int:
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return default


def safe_float(val, default: float = 0.0) -> float:
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default


# ─── 메시지 입력 ─────────────────────────────────────────
def type_message(text: str):
    """
    클립보드 방식으로 메시지 입력 (한글/이모지/특수문자 완벽 지원)
    기존 클립보드 내용 백업 후 복원
    """
    try:
        original = pyperclip.paste()
    except Exception:
        original = ""

    pyperclip.copy(text)
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.15)

    try:
        pyperclip.copy(original)
    except Exception:
        pass


# ─── 딜레이 ──────────────────────────────────────────────
def jitter_sleep(min_sec: float, max_sec: float):
    """랜덤 딜레이 (탐지 회피)"""
    delay = random.uniform(
        max(0.0, min_sec),
        max(0.0, max_sec)
    )
    time.sleep(delay)


def safe_sleep(seconds: float):
    """안전한 sleep (음수 방지)"""
    if seconds > 0:
        time.sleep(seconds)


# ─── Chrome 주소창 URL 입력 ──────────────────────────────
def navigate_chrome(url: str, addr_x: int, addr_y: int, load_delay: float):
    """
    Chrome 주소창에 URL 입력 후 Enter
    """
    pyautogui.click(addr_x, addr_y)
    time.sleep(0.3)

    pyautogui.hotkey('ctrl', 'a')
    time.sleep(0.1)

    pyperclip.copy(url)
    time.sleep(0.1)
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.2)

    keyboard.press_and_release('enter')
    safe_sleep(load_delay)


# ─── 좌표 Pick ───────────────────────────────────────────
def pick_coordinate(callback, countdown: int = 3):
    """
    카운트다운 후 마우스 현재 위치 캡처
    callback(result, mode):
        mode="countdown" → result = 남은 초 (int)
        mode="done"      → result = (x, y) tuple
        mode="error"     → result = 에러 메시지
    """
    def _run():
        try:
            for i in range(countdown, 0, -1):
                callback(i, mode="countdown")
                time.sleep(1)
            x, y = pyautogui.position()
            callback((x, y), mode="done")
        except Exception as e:
            callback(str(e), mode="error")

    t = threading.Thread(target=_run, daemon=True)
    t.start()


# ─── 날짜/시간 ───────────────────────────────────────────
def now_str() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def today_str() -> str:
    return datetime.now().strftime("%Y-%m-%d")


# ─── 로그 포맷 ───────────────────────────────────────────
def fmt_log(msg: str) -> str:
    ts = datetime.now().strftime("%H:%M:%S")
    return f"[{ts}] {msg}"


# ─── 플레이스홀더 치환 ────────────────────────────────────
def build_message(template: str, custom_msg: str = "", link: str = "") -> str:
    """
    {{custom_msg}} → custom_msg 값
    {{date}}       → 오늘 날짜
    {{link}}       → 그룹 링크
    """
    result = template
    result = result.replace("{{custom_msg}}", custom_msg or "")
    result = result.replace("{{date}}", today_str())
    result = result.replace("{{link}}", link or "")
    return result.strip()


# ─── 가입 상태 픽셀 감지 (10x10 영역 샘플링) ──────────────
def _is_blue(r: int, g: int, b: int) -> bool:
    """단일 픽셀이 텔레그램 JOIN 버튼 파란색 계열인지 판별"""
    # 텔레그램 파란색: B가 가장 높고, B-R 차이가 큼
    return (b > 150) and (b - r > 100) and (g > 80)


def has_blue_in_region(cx: int, cy: int, size: int = 10) -> bool:
    """
    (cx, cy) 중심으로 size x size 영역을 샘플링해
    파란색 픽셀이 1개라도 있으면 True 반환.

    파란색이 있다 = JOIN GROUP 버튼이 표시 중
    파란색이 없다 = Write a message 입력창으로 전환됨
    """
    try:
        shot = pyautogui.screenshot()
        half = size // 2
        x0 = max(0, cx - half)
        y0 = max(0, cy - half)
        x1 = cx + half
        y1 = cy + half

        for x in range(x0, x1):
            for y in range(y0, y1):
                try:
                    r, g, b = shot.getpixel((x, y))
                    if _is_blue(r, g, b):
                        return True
                except Exception:
                    continue
        return False
    except Exception:
        return False


def check_join_result(
    join_x: int,
    join_y: int,
    wait_sec: float = 2.5,
    retry: int = 3,
    retry_interval: float = 1.0,
) -> str:
    """
    가입 버튼 클릭 전후 10x10 영역 파란색 유무로 판정

    클릭 전  파란색 있음 → JOIN 버튼 존재
    클릭 전  파란색 없음 → 이미 가입된 채널 (Write a message 상태)
    클릭 후  파란색 없음 → 가입 완료
    클릭 후  파란색 있음 → 가입 실패 (버튼 그대로)

    반환: "가입완료" | "이미가입" | "실패"
    """
    safe_sleep(wait_sec)

    for attempt in range(retry):
        blue_found = has_blue_in_region(join_x, join_y, size=10)

        if not blue_found:
            return "가입완료"

        # 파란색 여전히 있음 → 실패, 재시도
        if attempt < retry - 1:
            safe_sleep(retry_interval)

    return "실패"


def check_already_joined(join_x: int, join_y: int) -> bool:
    """
    클릭 전 사전 체크: 이미 가입된 채널인지 확인
    파란색 없음 = 이미 가입 (Write a message 상태)
    """
    return not has_blue_in_region(join_x, join_y, size=10)


# ── A3: stop_manager.py ─────────────────────────────
# [DUP_REMOVED] import threading
# [DUP_CLEAN] from typing import Dict


class StopController:
    """
    탭별 독립적인 threading.Event 관리
    join_worker, message_worker 각각 별도 이벤트
    """

    def __init__(self):
        self._events: Dict[str, threading.Event] = {}
        self._lock = threading.Lock()

    def get_event(self, key: str) -> threading.Event:
        """키에 해당하는 이벤트 반환 (없으면 생성)"""
        with self._lock:
            if key not in self._events:
                self._events[key] = threading.Event()
            return self._events[key]

    def stop(self, key: str):
        """특정 작업 중단 신호"""
        with self._lock:
            if key in self._events:
                self._events[key].set()

    def stop_all(self):
        """모든 작업 중단"""
        with self._lock:
            for event in self._events.values():
                event.set()

    def reset(self, key: str):
        """작업 완료/재시작 전 초기화"""
        with self._lock:
            if key in self._events:
                self._events[key].clear()
            else:
                self._events[key] = threading.Event()

    def is_stopped(self, key: str) -> bool:
        """중단 여부 확인"""
        with self._lock:
            if key not in self._events:
                return False
            return self._events[key].is_set()


# 전역 싱글톤
stop_controller = StopController()


# ════════════════════════════════════════════════════════════
# [B] Data Layer — LinkParser / ExcelManager / ImageUtils
# ════════════════════════════════════════════════════════════

# ── B1: link_parser.py ──────────────────────────────
# [DUP_REMOVED] import re
# [DUP_CLEAN] from typing import List


# ─── 정규식 패턴 ─────────────────────────────────────────
_PATTERNS = [
    # https://t.me/xxx 또는 http://t.me/xxx
    r'https?://t\.me/[^\s\]\[\)\(\"\'<>]+',
    # t.me/xxx (프로토콜 없음)
    r'(?<![/\w])t\.me/[^\s\]\[\)\(\"\'<>]+',
    # @username
    r'(?<![/\w@])@[a-zA-Z][a-zA-Z0-9_]{3,}',
]


def parse_links(raw_text: str) -> List[str]:
    """
    텍스트에서 텔레그램 링크 추출
    다양한 형식 지원:
      - https://t.me/groupname
      - http://t.me/groupname
      - t.me/groupname
      - @groupname
      - https://t.me/+inviteHash
      - https://t.me/joinchat/XXXXX
    """
    if not raw_text:
        return []

    found = []
    for pattern in _PATTERNS:
        matches = re.findall(pattern, raw_text, re.IGNORECASE)
        found.extend(matches)

    # 정규화
    normalized = [normalize_link(lk) for lk in found]

    # None 제거 + 중복제거
    cleaned = [lk for lk in normalized if lk]
    return deduplicate(cleaned)


def normalize_link(link: str) -> str:
    """
    다양한 형식 → https://t.me/xxx 통일
    """
    if not link:
        return ""

    link = link.strip().rstrip('/')

    # @username → https://t.me/username
    if link.startswith('@'):
        username = link[1:]
        if _is_valid_username(username):
            return f"https://t.me/{username}"
        return ""

    # t.me/xxx → https://t.me/xxx
    if link.startswith('t.me/'):
        return "https://" + link

    # http:// → https://
    if link.startswith('http://t.me/'):
        return "https://t.me/" + link[len('http://t.me/'):]

    # 이미 https://t.me/
    if link.startswith('https://t.me/'):
        return link

    return ""


def deduplicate(links: List[str]) -> List[str]:
    """
    대소문자 무시 중복 제거 (순서 유지)
    https://t.me/+xxx 형태의 초대링크는 그대로 유지
    """
    seen = set()
    result = []
    for link in links:
        # 초대링크는 케이스 그대로, 일반링크는 소문자 비교
        if '/+' in link or '/joinchat/' in link.lower():
            key = link
        else:
            key = link.lower()

        if key not in seen:
            seen.add(key)
            result.append(link)
    return result


def validate_link(link: str) -> bool:
    """유효한 t.me 링크인지 검증"""
    if not link:
        return False
    if not link.startswith('https://t.me/'):
        return False
    path = link[len('https://t.me/'):]
    if not path:
        return False
    return True


def _is_valid_username(username: str) -> bool:
    """텔레그램 username 유효성 검사 (5자 이상, 영숫자+_)"""
    if len(username) < 4:
        return False
    return bool(re.match(r'^[a-zA-Z][a-zA-Z0-9_]{3,}$', username))


def extract_username(link: str) -> str:
    """
    https://t.me/groupname → groupname
    https://t.me/+hash → +hash
    """
    if not link.startswith('https://t.me/'):
        return link
    return link[len('https://t.me/'):]


def count_stats(raw_text: str) -> dict:
    """
    파싱 통계 반환
    returns: {"raw": N, "parsed": M, "duplicates": K}
    """
    # 줄 수 (raw 입력)
    raw_lines = [l.strip() for l in raw_text.splitlines() if l.strip()]

    # 파싱된 링크
    parsed = parse_links(raw_text)

    # 전체 추출 (중복 포함)
    all_found = []
    for pattern in _PATTERNS:
        matches = re.findall(pattern, raw_text, re.IGNORECASE)
        normalized = [normalize_link(m) for m in matches]
        all_found.extend([n for n in normalized if n])

    return {
        "raw": len(raw_lines),
        "parsed": len(parsed),
        "duplicates": len(all_found) - len(parsed),
    }


# ── B2: excel_manager.py ─────────────────────────────
# [DUP_REMOVED] import os
# [DUP_REMOVED] from dataclasses import dataclass
# [DUP_CLEAN] from typing import List, Optional
# [DUP_REMOVED] from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# [MERGED] import config as cfg


@dataclass
class GroupRow:
    """그룹 정보 데이터 클래스"""
    row_index: int          # Excel 행 번호 (2-based, 1행은 헤더)
    link: str               # https://t.me/xxx
    status: str = "미가입"  # 미가입/가입완료/발송완료/실패/건너뜀
    custom_msg: str = ""    # 커스텀 메시지
    image_path: str = ""    # 이미지 경로
    processed_at: str = ""  # 처리일시
    note: str = ""          # 비고


class ExcelManager:
    """
    Excel 파일 로드/저장/실시간 업데이트
    """

    def __init__(self, filepath: str = ""):
        self.filepath = filepath
        self._rows: List[GroupRow] = []

    # ─── 로드 ────────────────────────────────────────────
    def load(self, filepath: str = "") -> List[GroupRow]:
        """Excel 파일 읽기 → GroupRow 리스트 반환"""
        if filepath:
            self.filepath = filepath

        if not self.filepath or not os.path.exists(self.filepath):
            raise FileNotFoundError(f"파일을 찾을 수 없습니다: {self.filepath}")

        if not HAS_OPENPYXL:
            raise ImportError("openpyxl이 설치되지 않았습니다. pip install openpyxl")

        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            link = str(row[0]).strip() if row[0] else ""
            if not link:
                continue

            rows.append(GroupRow(
                row_index=i,
                link=link,
                status=str(row[1]).strip() if len(row) > 1 and row[1] else "미가입",
                custom_msg=str(row[2]).strip() if len(row) > 2 and row[2] else "",
                image_path=str(row[3]).strip() if len(row) > 3 and row[3] else "",
                processed_at=str(row[4]).strip() if len(row) > 4 and row[4] else "",
                note=str(row[5]).strip() if len(row) > 5 and row[5] else "",
            ))

        self._rows = rows
        wb.close()
        return rows

    # ─── 저장 ────────────────────────────────────────────
    def save_all(self, rows: List[GroupRow] = None, filepath: str = ""):
        """전체 저장"""
        if rows is None:
            rows = self._rows
        if filepath:
            self.filepath = filepath
        if not self.filepath:
            raise ValueError("저장 경로가 없습니다.")

        if not HAS_OPENPYXL:
            raise ImportError("openpyxl이 설치되지 않았습니다.")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "그룹목록"

        # 헤더
        self._write_header(ws)

        # 데이터
        for i, row in enumerate(rows, start=2):
            ws.cell(i, EXCEL_COL["link"],       row.link)
            ws.cell(i, EXCEL_COL["status"],     row.status)
            ws.cell(i, EXCEL_COL["custom_msg"], row.custom_msg)
            ws.cell(i, EXCEL_COL["image_path"], row.image_path)
            ws.cell(i, EXCEL_COL["processed"],  row.processed_at)
            ws.cell(i, EXCEL_COL["note"],       row.note)

        # 열 너비 조정
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        wb.save(self.filepath)
        wb.close()

    def update_row(self, row_index: int, status: str, note: str = ""):
        """
        단일 행 실시간 업데이트
        - 작업 중 즉시 저장 → 강제종료 되어도 진행상황 보존
        """
        if not self.filepath or not os.path.exists(self.filepath):
            return

        try:
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb.active

            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row_index, EXCEL_COL["status"]).value = status
            ws.cell(row_index, EXCEL_COL["processed"]).value = now
            if note:
                ws.cell(row_index, EXCEL_COL["note"]).value = note

            wb.save(self.filepath)
            wb.close()

            # 메모리 rows도 업데이트
            for r in self._rows:
                if r.row_index == row_index:
                    r.status = status
                    r.processed_at = now
                    if note:
                        r.note = note
                    break

        except Exception:
            pass  # 파일 잠금 등 무시

    # ─── 신규 생성 ────────────────────────────────────────
    def create_from_links(self, links: List[str], filepath: str):
        """링크 목록 → 신규 Excel 생성 (모두 미가입 상태)"""
        self.filepath = filepath
        rows = [
            GroupRow(row_index=i + 2, link=lk)
            for i, lk in enumerate(links)
        ]
        self._rows = rows
        self.save_all(rows)
        return rows

    # ─── 필터링 ──────────────────────────────────────────
    def get_filtered(self, status_filter: str) -> List[GroupRow]:
        """상태별 필터링"""
        if status_filter == "전체":
            return self._rows[:]
        elif status_filter == "미가입만":
            return [r for r in self._rows if r.status == "미가입"]
        elif status_filter == "가입완료만":
            return [r for r in self._rows if r.status == "가입완료"]
        elif status_filter == "발송완료만":
            return [r for r in self._rows if r.status == "발송완료"]
        elif status_filter == "실패만":
            return [r for r in self._rows if r.status == "실패"]
        elif status_filter == "재발송 포함":
            return [r for r in self._rows
                    if r.status in ("가입완료", "발송완료")]
        else:
            return self._rows[:]

    def get_stats(self) -> dict:
        """상태 통계"""
        stats = {
            "전체": len(self._rows),
            "미가입": 0,
            "가입완료": 0,
            "발송완료": 0,
            "실패": 0,
            "건너뜀": 0,
        }
        for r in self._rows:
            if r.status in stats:
                stats[r.status] += 1
        return stats

    @property
    def rows(self) -> List[GroupRow]:
        return self._rows

    # ─── 내부 헬퍼 ───────────────────────────────────────
    def _write_header(self, ws):
        """헤더 행 작성 (스타일 포함)"""
        header_fill = PatternFill("solid", fgColor="0F3460")
        header_font = Font(color="EAEAEA", bold=True, name="맑은 고딕", size=10)

        for col_idx, header in enumerate(EXCEL_HEADERS, start=1):
            cell = ws.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")


# ── B3: image_utils.py ───────────────────────────────
# [DUP_REMOVED] import os
# [DUP_REMOVED] import time
# [DUP_REMOVED] import pyautogui
# [DUP_REMOVED] import pyperclip
# [DUP_REMOVED] import keyboard
# [DUP_CLEAN] from typing import Optional

SUPPORTED_EXTS = {'.png', '.jpg', '.jpeg', '.gi', '.mp4', '.webp', '.bmp'}


def validate_image_path(path: str) -> bool:
    """파일 존재 여부 + 지원 형식 확인"""
    if not path:
        return False
    if not os.path.exists(path):
        return False
    ext = os.path.splitext(path)[1].lower()
    return ext in SUPPORTED_EXTS


def get_image_for_row(image_path_from_excel: str, default_folder: str) -> Optional[str]:
    """
    이미지 경로 결정 우선순위:
    1. Excel D열 개별 경로
    2. default_folder 의 첫 번째 이미지
    3. None
    """
    # 1순위: Excel 개별 경로
    if image_path_from_excel and validate_image_path(image_path_from_excel):
        return image_path_from_excel

    # 2순위: 기본 폴더
    if default_folder and os.path.isdir(default_folder):
        for fname in sorted(os.listdir(default_folder)):
            ext = os.path.splitext(fname)[1].lower()
            if ext in SUPPORTED_EXTS:
                return os.path.join(default_folder, fname)

    return None


def attach_image_via_dialog(
    image_path: str,
    attach_btn_x: int,
    attach_btn_y: int,
    file_addr_x: int,
    file_addr_y: int,
    delays: dict,
    cancel_event=None,
    logger=None,
) -> bool:
    """
    파일 다이얼로그 방식 이미지 첨부
    흐름:
      ① 📎 첨부버튼 클릭
      ② 파일탐색기 열림 대기
      ③ 주소창 클릭
      ④ 폴더 경로 입력 + Enter
      ⑤ 파일명 선택 (파일명 타이핑)
      ⑥ 열기/Enter
    """

    def _log(msg):
        if logger:
            logger(msg)

    def _cancelled():
        return cancel_event is not None and cancel_event.is_set()

    if not validate_image_path(image_path):
        _log(f"⚠️ 이미지 파일 없음: {image_path}")
        return False

    folder = os.path.dirname(image_path)
    filename = os.path.basename(image_path)

    try:
        # ① 첨부버튼 클릭
        pyautogui.click(attach_btn_x, attach_btn_y)
        time.sleep(delays.get("file_dialog_open", 1.5))
        if _cancelled():
            return False

        # ② 파일탐색기 주소창 클릭
        pyautogui.click(file_addr_x, file_addr_y)
        time.sleep(delays.get("after_addr_click", 0.5))
        if _cancelled():
            return False

        # ③ 폴더 경로 입력
        pyautogui.hotkey('ctrl', 'a')
        time.sleep(0.1)
        pyperclip.copy(folder)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.1)
        keyboard.press_and_release('enter')
        time.sleep(delays.get("after_path_input", 0.8))
        if _cancelled():
            return False

        # ④ 파일명 입력창 클릭 후 파일명 입력
        # 파일탐색기 하단 파일명 입력란 클릭 (파일탐색기 중앙 하단)
        # 파일명 입력란은 file_addr_y + 일정 오프셋 (사용자 좌표 설정)
        # → 여기서는 Ctrl+A 후 파일명 직접 입력
        pyperclip.copy(filename)
        pyautogui.hotkey('ctrl', 'v')
        time.sleep(0.2)
        keyboard.press_and_release('enter')
        time.sleep(delays.get("after_file_enter", 2.0))

        _log(f"📎 이미지 첨부: {filename}")
        return True

    except Exception as e:
        _log(f"❌ 이미지 첨부 오류: {e}")
        return False


def get_files_in_folder(folder: str) -> list:
    """폴더 내 지원 이미지 파일 목록"""
    if not folder or not os.path.isdir(folder):
        return []
    result = []
    for fname in sorted(os.listdir(folder)):
        ext = os.path.splitext(fname)[1].lower()
        if ext in SUPPORTED_EXTS:
            result.append(os.path.join(folder, fname))
    return result


# ════════════════════════════════════════════════════════════
# [C] Worker Layer — JoinWorker / MessageWorker
# ════════════════════════════════════════════════════════════

# ── C1: join_worker.py ──────────────────────────────
# [DUP_REMOVED] import time
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import pyautogui
# [DUP_REMOVED] import keyboard

# [DUP_CLEAN] from utils.helpers import (
# [DUP_CLEAN]     navigate_chrome, jitter_sleep, safe_sleep, fmt_log,
# [DUP_CLEAN]     has_blue_in_region, check_join_result, check_already_joined
# [DUP_CLEAN] )
# [DUP_CLEAN] from utils.image_utils import attach_image_via_dialog, get_image_for_row
# [DUP_CLEAN] from utils.excel_manager import GroupRow, ExcelManager


class JoinWorker:
    """
    텔레그램 그룹 자동 합류 워커

    [판정 로직]
    1. Chrome에서 t.me 링크 열기
    2. "Telegram에서 열기" 클릭
    3. JOIN 버튼 좌표의 픽셀 색상 읽기 (클릭 전)
       - 이미 다크 → "이미가입" 처리 (클릭 스킵)
       - 파란색    → JOIN 버튼 클릭
    4. 클릭 후 2~3초 대기, 같은 좌표 재확인
       - 다크로 바뀜 → "가입완료"
       - 여전히 파란 → "실패"
    """

    def __init__(self):
        self._running = False

    def run(
        self,
        rows: list,
        config: dict,
        cancel_event: threading.Event,
        logger,
        excel_manager: ExcelManager,
        progress_callback=None,
    ):
        self._running = True
        total = len(rows)
        success = 0
        fail = 0

        send_after_join = config.get("send_after_join", False)
        join_msg = config.get("join_msg", "")

        # 픽셀 체크 대기 시간 (coord_tab에서 설정, 기본 2.5초)
        check_delay  = config.get("join_result_check_delay", 2.5)
        check_retry  = int(config.get("join_result_check_retry", 3))

        for idx, row in enumerate(rows):
            if self._should_stop(cancel_event):
                logger(fmt_log("⏹ 중단됨"))
                break

            logger(fmt_log(f"처리 중 [{idx+1}/{total}]: {row.link}"))

            try:
                result = self._process_one(
                    row=row,
                    config=config,
                    cancel_event=cancel_event,
                    logger=logger,
                    excel_manager=excel_manager,
                    send_after_join=send_after_join,
                    join_msg=join_msg,
                    check_delay=check_delay,
                    check_retry=check_retry,
                )

                if result in ("가입완료", "이미가입"):
                    success += 1
                else:
                    fail += 1

            except Exception as e:
                excel_manager.update_row(row.row_index, "실패", str(e))
                logger(fmt_log(f"❌ {row.link} → 오류: {e}"))
                fail += 1

            if progress_callback:
                progress_callback(idx + 1, total, success, fail)

            # 링크 간 랜덤 딜레이
            if not self._should_stop(cancel_event) and idx < total - 1:
                jitter_sleep(
                    config.get("between_min", 3.0),
                    config.get("between_max", 7.0),
                )

        self._running = False
        logger(fmt_log(f"🏁 완료 — 성공: {success}, 실패: {fail}"))

    # ─── 링크 1개 처리 ───────────────────────────────────
    def _process_one(
        self, row, config, cancel_event, logger,
        excel_manager, send_after_join, join_msg,
        check_delay, check_retry,
    ) -> str:
        """
        반환: "가입완료" | "이미가입" | "실패" | "알수없음"
        """
        join_x = config["join_btn_x"]
        join_y = config["join_btn_y"]

        # ── Step 1: Chrome 주소창 열기 ──────────────────
        navigate_chrome(
            url=row.link,
            addr_x=config["chrome_addr_x"],
            addr_y=config["chrome_addr_y"],
            load_delay=config.get("chrome_load_delay", 2.0),
        )
        if self._should_stop(cancel_event):
            return "알수없음"

        # ── Step 2: "Telegram에서 열기" 클릭 ────────────
        pyautogui.click(
            config["chrome_open_btn_x"],
            config["chrome_open_btn_y"]
        )
        safe_sleep(config.get("telegram_open_delay", 1.5))
        if self._should_stop(cancel_event):
            return "알수없음"

        # ── Step 3: 클릭 전 사전 체크 (이미 가입 여부) ──
        if check_already_joined(join_x, join_y):
            logger(fmt_log("  ℹ️ 이미 가입된 채널 — 클릭 스킵"))
            excel_manager.update_row(row.row_index, "가입완료", "이미가입")
            pyautogui.click(config["back_btn_x"], config["back_btn_y"])
            safe_sleep(config.get("after_back_delay", 0.8))
            return "이미가입"

        # ── Step 4: JOIN GROUP 버튼 클릭 ────────────────
        pyautogui.click(join_x, join_y)
        logger(fmt_log("  🖱️ JOIN GROUP 클릭 → 10x10 영역 파란색 확인 중..."))

        # ── Step 5: 10x10 영역 파란색 유무로 판정 ────────
        result = check_join_result(
            join_x=join_x,
            join_y=join_y,
            wait_sec=check_delay,
            retry=check_retry,
            retry_interval=1.0,
        )

        if self._should_stop(cancel_event):
            return "알수없음"

        # ── Step 6: 결과에 따른 처리 ────────────────────
        if result == "가입완료":
            if send_after_join and join_msg:
                # 통합 모드: 합류 후 즉시 메시지
                self._send_message_after_join(
                    row=row, msg=join_msg,
                    config=config,
                    cancel_event=cancel_event,
                    logger=logger,
                )
                if not self._should_stop(cancel_event):
                    excel_manager.update_row(row.row_index, "발송완료")
                    logger(fmt_log(f"✅ {row.link} → 가입 + 발송 완료"))
            else:
                # 뒤로가기
                pyautogui.click(config["back_btn_x"], config["back_btn_y"])
                safe_sleep(config.get("after_back_delay", 0.8))
                excel_manager.update_row(row.row_index, "가입완료")
                logger(fmt_log(f"✅ {row.link} → 가입 완료"))

        elif result == "이미가입":
            pyautogui.click(config["back_btn_x"], config["back_btn_y"])
            safe_sleep(config.get("after_back_delay", 0.8))
            excel_manager.update_row(row.row_index, "가입완료", "이미가입")
            logger(fmt_log(f"ℹ️ {row.link} → 이미 가입된 채널"))

        elif result == "실패":
            pyautogui.click(config["back_btn_x"], config["back_btn_y"])
            safe_sleep(config.get("after_back_delay", 0.8))
            excel_manager.update_row(row.row_index, "실패", "JOIN버튼 변화없음")
            logger(fmt_log(f"❌ {row.link} → 가입 실패 (버튼 변화 없음)"))

        else:  # 알수없음
            pyautogui.click(config["back_btn_x"], config["back_btn_y"])
            safe_sleep(config.get("after_back_delay", 0.8))
            excel_manager.update_row(row.row_index, "실패", "픽셀판정불가")
            logger(fmt_log(f"⚠️ {row.link} → 판정 불가 (실패로 기록)"))

        return result

    # ─── 합류 후 메시지 발송 ─────────────────────────────
    def _send_message_after_join(
        self, row: GroupRow, msg: str, config: dict,
        cancel_event, logger
    ):
        """통합 모드: 합류 직후 메시지 발송"""
        from utils.helpers import type_message, build_message

        final_msg = build_message(msg, row.custom_msg, row.link)

        # 이미지 before
        if (config.get("image_mode") == "files"
                and config.get("image_timing") == "before"):
            img = get_image_for_row(row.image_path, config.get("image_folder", ""))
            if img:
                attach_image_via_dialog(
                    img,
                    config["attach_btn_x"], config["attach_btn_y"],
                    config["file_addr_x"], config["file_addr_y"],
                    config, cancel_event, logger,
                )

        if self._should_stop(cancel_event):
            return

        # 메시지 입력
        if final_msg:
            pyautogui.click(config["msg_input_x"], config["msg_input_y"])
            safe_sleep(0.3)
            type_message(final_msg)
            safe_sleep(config.get("after_type_delay", 0.5))

        # 이미지 after
        if (config.get("image_mode") == "files"
                and config.get("image_timing") == "after"):
            img = get_image_for_row(row.image_path, config.get("image_folder", ""))
            if img:
                attach_image_via_dialog(
                    img,
                    config["attach_btn_x"], config["attach_btn_y"],
                    config["file_addr_x"], config["file_addr_y"],
                    config, cancel_event, logger,
                )

        if self._should_stop(cancel_event):
            return

        # 전송
        if config.get("send_method", "enter") == "enter":
            keyboard.press_and_release('enter')
        else:
            pyautogui.click(config["send_btn_x"], config["send_btn_y"])
        safe_sleep(config.get("after_send_delay", 1.0))

        # 뒤로가기
        pyautogui.click(config["back_btn_x"], config["back_btn_y"])
        safe_sleep(config.get("after_back_delay", 0.8))

    def _should_stop(self, cancel_event: threading.Event) -> bool:
        return (cancel_event is not None and cancel_event.is_set()) \
               or not self._running

    def stop(self):
        self._running = False


# ── C2: message_worker.py ───────────────────────────
# [DUP_REMOVED] import time
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import pyautogui
# [DUP_REMOVED] import keyboard

# [DUP_CLEAN] from utils.helpers import (
# [DUP_CLEAN]     navigate_chrome, type_message, build_message,
# [DUP_CLEAN]     jitter_sleep, safe_sleep, fmt_log
# [DUP_CLEAN] )
# [DUP_CLEAN] from utils.image_utils import attach_image_via_dialog, get_image_for_row
# [DUP_CLEAN] from utils.excel_manager import GroupRow, ExcelManager


class MessageWorker:
    """
    텔레그램 그룹 자동 메시지 발송 워커
    - Chrome으로 t.me 링크 열기
    - "Telegram에서 열기" 버튼 클릭
    - 이미지 첨부 (선택)
    - 메시지 입력 + 전송
    - 결과 Excel 기록
    """

    def __init__(self):
        self._running = False

    def run(
        self,
        rows: list,
        config: dict,
        cancel_event: threading.Event,
        logger,
        excel_manager: ExcelManager,
        progress_callback=None,
    ):
        """
        메인 실행 루프
        rows: List[GroupRow]
        config: 좌표 + 딜레이 + 메시지 설정 dict
        """
        self._running = True
        total = len(rows)
        success = 0
        fail = 0

        default_msg = config.get("default_msg", "")
        image_mode = config.get("image_mode", "none")
        image_timing = config.get("image_timing", "after")
        image_folder = config.get("image_folder", "")

        for idx, row in enumerate(rows):
            if self._should_stop(cancel_event):
                logger(fmt_log("⏹ 중단됨"))
                break

            logger(fmt_log(f"처리 중 [{idx+1}/{total}]: {row.link}"))

            # 최종 메시지 완성
            final_msg = build_message(default_msg, row.custom_msg, row.link)

            # 이미지 경로 결정
            img_path = None
            if image_mode == "files":
                img_path = get_image_for_row(row.image_path, image_folder)

            try:
                # ── Step 1: Chrome 주소창 URL 입력 ──
                navigate_chrome(
                    url=row.link,
                    addr_x=config["chrome_addr_x"],
                    addr_y=config["chrome_addr_y"],
                    load_delay=config.get("chrome_load_delay", 2.0),
                )
                if self._should_stop(cancel_event):
                    break

                # ── Step 2: "Telegram에서 열기" 버튼 클릭 ──
                pyautogui.click(
                    config["chrome_open_btn_x"],
                    config["chrome_open_btn_y"]
                )
                safe_sleep(config.get("telegram_open_delay", 1.5))
                if self._should_stop(cancel_event):
                    break

                # ── Step 3A: 이미지 먼저 (timing=before) ──
                if image_mode == "files" and image_timing == "before" and img_path:
                    attach_image_via_dialog(
                        img_path,
                        config["attach_btn_x"], config["attach_btn_y"],
                        config["file_addr_x"], config["file_addr_y"],
                        config, cancel_event, logger,
                    )
                    if self._should_stop(cancel_event):
                        break

                # ── Step 3B: 메시지 입력 ──
                if final_msg:
                    pyautogui.click(config["msg_input_x"], config["msg_input_y"])
                    safe_sleep(0.3)
                    type_message(final_msg)
                    safe_sleep(config.get("after_type_delay", 0.5))
                    if self._should_stop(cancel_event):
                        break

                # ── Step 3C: 이미지 나중 (timing=after) ──
                if image_mode == "files" and image_timing == "after" and img_path:
                    attach_image_via_dialog(
                        img_path,
                        config["attach_btn_x"], config["attach_btn_y"],
                        config["file_addr_x"], config["file_addr_y"],
                        config, cancel_event, logger,
                    )
                    if self._should_stop(cancel_event):
                        break

                # ── Step 4: 전송 ──
                if config.get("send_method", "enter") == "enter":
                    keyboard.press_and_release('enter')
                else:
                    pyautogui.click(config["send_btn_x"], config["send_btn_y"])
                safe_sleep(config.get("after_send_delay", 1.0))
                if self._should_stop(cancel_event):
                    break

                # ── Step 5: 뒤로가기 ──
                pyautogui.click(config["back_btn_x"], config["back_btn_y"])
                safe_sleep(config.get("after_back_delay", 0.8))

                # ── Step 6: 결과 기록 ──
                excel_manager.update_row(row.row_index, "발송완료")
                logger(fmt_log(f"✅ {row.link} → 발송 완료"))
                success += 1

            except Exception as e:
                excel_manager.update_row(row.row_index, "실패", str(e))
                logger(fmt_log(f"❌ {row.link} → 오류: {e}"))
                fail += 1

            # 진행상황 콜백
            if progress_callback:
                progress_callback(idx + 1, total, success, fail)

            # ── Step 7: 링크 간 랜덤 딜레이 ──
            if not self._should_stop(cancel_event) and idx < total - 1:
                jitter_sleep(
                    config.get("between_min", 2.0),
                    config.get("between_max", 6.0),
                )

        self._running = False
        logger(fmt_log(f"🏁 완료 — 성공: {success}, 실패: {fail}"))

    def _should_stop(self, cancel_event: threading.Event) -> bool:
        return (cancel_event is not None and cancel_event.is_set()) \
               or not self._running

    def stop(self):
        self._running = False


# ════════════════════════════════════════════════════════════
# [D] UI Layer
# ════════════════════════════════════════════════════════════

# ── D1: widgets.py ──────────────────────────────────
# [DUP_REMOVED] import tkinter as tk
# [DUP_CLEAN] from tkinter import ttk
# [MERGED] import config as cfg


class ScrollableFrame(tk.Frame):
    """세로 스크롤 가능한 프레임"""

    def __init__(self, parent, **kwargs):
        bg = kwargs.pop("bg", THEME["card"])
        super().__init__(parent, bg=bg, **kwargs)

        # Canvas + Scrollbar
        self.canvas = tk.Canvas(self, bg=bg, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self, orient="vertical",
                                  command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)

        # 내부 프레임
        self.inner = tk.Frame(self.canvas, bg=bg)
        self._window_id = self.canvas.create_window(
            (0, 0), window=self.inner, anchor="nw"
        )

        # 이벤트 바인딩
        self.inner.bind("<Configure>", self._on_inner_configure)
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<MouseWheel>", self._on_mousewheel)
        self.inner.bind("<MouseWheel>", self._on_mousewheel)

    def _on_inner_configure(self, event):
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self.canvas.itemconfig(self._window_id, width=event.width)

    def _on_mousewheel(self, event):
        self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")


class HoverTooltip:
    """마우스 호버 시 툴팁 표시"""

    def __init__(self, widget, text: str, delay: int = 600):
        self.widget = widget
        self.text = text
        self.delay = delay
        self._id = None
        self._tip_window = None

        widget.bind("<Enter>", self._schedule)
        widget.bind("<Leave>", self._hide)
        widget.bind("<ButtonPress>", self._hide)

    def _schedule(self, event=None):
        self._unschedule()
        self._id = self.widget.after(self.delay, self._show)

    def _unschedule(self):
        if self._id:
            self.widget.after_cancel(self._id)
            self._id = None

    def _show(self, event=None):
        if self._tip_window:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 4

        self._tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")

        lbl = tk.Label(
            tw, text=self.text,
            bg="#1a1a2e", fg="#eaeaea",
            font=("맑은 고딕", 9),
            relief="solid", bd=1,
            padx=6, pady=4
        )
        lbl.pack()

    def _hide(self, event=None):
        self._unschedule()
        if self._tip_window:
            self._tip_window.destroy()
            self._tip_window = None


class LogText(tk.Frame):
    """
    로그 출력 텍스트박스
    append(msg) 메서드로 로그 추가
    """

    def __init__(self, parent, height: int = 10, **kwargs):
        bg = kwargs.pop("bg", THEME["bg"])
        super().__init__(parent, bg=bg)

        self.text = tk.Text(
            self,
            height=height,
            bg=THEME["input_bg"],
            fg=THEME["text"],
            font=FONT["mono"],
            state="disabled",
            wrap="word",
            relief="flat",
            bd=0,
        )
        sb = ttk.Scrollbar(self, command=self.text.yview)
        self.text.configure(yscrollcommand=sb.set)

        sb.pack(side="right", fill="y")
        self.text.pack(side="left", fill="both", expand=True)

        # 태그 색상
        self.text.tag_configure("success", foreground=THEME["success"])
        self.text.tag_configure("warning", foreground=THEME["warning"])
        self.text.tag_configure("error",   foreground=THEME["danger"])
        self.text.tag_configure("info",    foreground=THEME["subtext"])

    def append(self, msg: str):
        """로그 추가 (자동 스크롤)"""
        self.text.configure(state="normal")

        # 태그 결정
        tag = None
        if any(c in msg for c in ["✅", "완료", "성공"]):
            tag = "success"
        elif any(c in msg for c in ["⚠️", "경고"]):
            tag = "warning"
        elif any(c in msg for c in ["❌", "오류", "실패", "에러"]):
            tag = "error"
        elif any(c in msg for c in ["[", "처리 중", "⏹", "🏁"]):
            tag = "info"

        if tag:
            self.text.insert("end", msg + "\n", tag)
        else:
            self.text.insert("end", msg + "\n")

        # 최대 1000줄 유지
        lines = int(self.text.index("end-1c").split(".")[0])
        if lines > 1000:
            self.text.delete("1.0", f"{lines - 1000}.0")

        self.text.configure(state="disabled")
        self.text.see("end")

    def clear(self):
        self.text.configure(state="normal")
        self.text.delete("1.0", "end")
        self.text.configure(state="disabled")


class LabeledEntry(tk.Frame):
    """레이블 + 입력창 조합 위젯"""

    def __init__(self, parent, label: str, width: int = 8,
                 var=None, **kwargs):
        bg = kwargs.pop("bg", THEME["card"])
        super().__init__(parent, bg=bg)

        self.var = var or tk.StringVar()

        tk.Label(
            self, text=label,
            bg=bg, fg=THEME["subtext"],
            font=FONT["small"]
        ).pack(side="left", padx=(0, 4))

        self.entry = tk.Entry(
            self,
            textvariable=self.var,
            width=width,
            bg=THEME["input_bg"],
            fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["small"],
            relief="flat",
            bd=1,
        )
        self.entry.pack(side="left")

    def get(self) -> str:
        return self.var.get()

    def set(self, val):
        self.var.set(str(val))


def make_btn(parent, text: str, command=None,
             color: str = None, width: int = None, **kwargs) -> tk.Button:
    """통일된 스타일의 버튼 생성"""
    bg = color or THEME["accent"]
    btn_kwargs = dict(
        text=text,
        command=command,
        bg=bg,
        fg="#ffffff",
        activebackground=bg,
        activeforeground="#ffffff",
        font=FONT["btn"],
        relief="flat",
        bd=0,
        padx=10,
        pady=5,
        cursor="hand2",
    )
    if width:
        btn_kwargs["width"] = width
    btn_kwargs.update(kwargs)
    return tk.Button(parent, **btn_kwargs)


def make_label(parent, text: str, color: str = None,
               font=None, **kwargs) -> tk.Label:
    """통일된 스타일의 레이블 생성"""
    lbl_kwargs = dict(
        text=text,
        bg=kwargs.pop("bg", THEME["card"]),
        fg=color or THEME["text"],
        font=font or FONT["label"],
    )
    lbl_kwargs.update(kwargs)
    return tk.Label(parent, **lbl_kwargs)


def make_section(parent, title: str) -> tk.LabelFrame:
    """섹션 구분용 LabelFrame"""
    return tk.LabelFrame(
        parent,
        text=f"  {title}  ",
        bg=THEME["card"],
        fg=THEME["accent"],
        font=FONT["header"],
        relief="groove",
        bd=1,
        labelanchor="nw",
    )


# ── D2: coord_tab.py ─────────────────────────────────
# [DUP_REMOVED] import json
# [DUP_REMOVED] import os
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import tkinter as tk
# [DUP_CLEAN] from tkinter import ttk, messagebox, simpledialog

# [MERGED] import config as cfg
# [DUP_CLEAN] from ui.widgets import (
# [DUP_CLEAN]     ScrollableFrame, LogText, make_btn, make_label,
# [DUP_CLEAN]     make_section, LabeledEntry
# [DUP_CLEAN] )
# [DUP_CLEAN] from utils.helpers import pick_coordinate, safe_float, safe_int, fmt_log


# ─── 좌표 필드 정의 ──────────────────────────────────────
COORD_FIELDS = [
    # (key_x, key_y, 표시이름, 설명)
    ("chrome_addr_x",     "chrome_addr_y",     "Chrome 주소창",
     "Chrome 브라우저 상단 주소창"),
    ("chrome_open_btn_x", "chrome_open_btn_y", "Telegram에서 열기 버튼",
     "t.me 페이지의 '열기' 버튼 위치"),
    ("join_btn_x",        "join_btn_y",         "가입하기 버튼",
     "텔레그램 앱의 '가입하기' 버튼"),
    ("msg_input_x",       "msg_input_y",        "메시지 입력창",
     "텔레그램 앱 메시지 입력 필드"),
    ("send_btn_x",        "send_btn_y",         "전송 버튼",
     "텔레그램 앱 전송 버튼"),
    ("attach_btn_x",      "attach_btn_y",       "첨부 버튼 (📎)",
     "텔레그램 앱 파일 첨부 버튼"),
    ("file_addr_x",       "file_addr_y",        "파일탐색기 주소창",
     "파일탐색기 상단 경로 입력창"),
    ("back_btn_x",        "back_btn_y",         "뒤로가기 버튼",
     "텔레그램 앱 뒤로가기 또는 닫기"),
]

# ─── 딜레이 필드 정의 ────────────────────────────────────
DELAY_FIELDS = [
    ("chrome_load_delay",           "Chrome 로딩 대기",       "Chrome에서 t.me 페이지 로드 후 대기"),
    ("telegram_open_delay",         "Telegram 전환 대기",     "'열기' 클릭 후 앱 전환 대기"),
    ("join_click_delay",            "가입 후 대기",            "가입버튼 클릭 후 대기"),
    ("join_result_check_delay",     "가입판정 대기",           "JOIN 클릭 후 픽셀 변화 확인 전 대기 (기본 2.5초)"),
    ("join_result_check_retry",     "가입판정 재시도 횟수",    "픽셀 판정 최대 재시도 횟수 (기본 3회, 정수)"),
    ("after_type_delay",            "타이핑 후 대기",          "메시지 입력 완료 후 대기"),
    ("after_send_delay",            "전송 후 대기",            "메시지 전송 후 대기"),
    ("after_back_delay",            "뒤로가기 후 대기",        "뒤로가기 후 다음 동작 전 대기"),
    ("between_min",                 "링크간 딜레이 최소",      "다음 링크 처리 전 최소 대기 (랜덤)"),
    ("between_max",                 "링크간 딜레이 최대",      "다음 링크 처리 전 최대 대기 (랜덤)"),
    ("file_dialog_open",            "파일탐색기 열림 대기",    "첨부버튼 클릭 후 파일탐색기 열림 대기"),
    ("after_addr_click",            "주소창 클릭 후 대기",     "파일탐색기 주소창 클릭 후 대기"),
    ("after_path_input",            "경로 입력 후 대기",       "폴더 경로 입력 후 대기"),
    ("after_file_enter",            "파일 선택 후 대기",       "파일 선택 후 업로드 완료 대기"),
]


class CoordTab(tk.Frame):
    """📍 좌표 설정 탭"""

    def __init__(self, parent):
        super().__init__(parent, bg=THEME["bg"])
        self._coord_vars = {}   # key → (var_x, var_y)
        self._delay_vars = {}   # key → var
        self._picking = False

        self._build_ui()
        self._load_config()

    # ─── UI 구성 ─────────────────────────────────────────
    def _build_ui(self):
        # 상단 프리셋 바
        self._build_preset_bar()

        # 좌우 분할
        paned = tk.PanedWindow(
            self, orient="horizontal",
            bg=THEME["bg"], sashwidth=4
        )
        paned.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        # 왼쪽: 좌표 설정
        left = tk.Frame(paned, bg=THEME["bg"])
        paned.add(left, minsize=480)
        self._build_coord_section(left)

        # 오른쪽: 딜레이 + 로그
        right = tk.Frame(paned, bg=THEME["bg"])
        paned.add(right, minsize=300)
        self._build_delay_section(right)
        self._build_log_section(right)

    def _build_preset_bar(self):
        bar = tk.Frame(self, bg=THEME["card"], height=44)
        bar.pack(fill="x", padx=8, pady=(8, 0))
        bar.pack_propagate(False)

        make_label(bar, "프리셋:", bg=THEME["card"]).pack(
            side="left", padx=(12, 4), pady=8)

        self._preset_var = tk.StringVar(value="기본값")
        self._preset_combo = ttk.Combobox(
            bar, textvariable=self._preset_var,
            width=16, state="readonly"
        )
        self._preset_combo.pack(side="left", padx=4, pady=8)
        self._preset_combo.bind("<<ComboboxSelected>>", self._on_preset_select)

        make_btn(bar, "📂 불러오기",  self._load_preset,
                 color=THEME["border"]).pack(side="left", padx=4, pady=6)
        make_btn(bar, "💾 저장",      self._save_preset,
                 color=THEME["accent"]).pack(side="left", padx=4, pady=6)
        make_btn(bar, "➕ 새 프리셋", self._new_preset,
                 color=THEME["success"]).pack(side="left", padx=4, pady=6)
        make_btn(bar, "🗑️ 삭제",     self._delete_preset,
                 color=THEME["danger"]).pack(side="left", padx=4, pady=6)

    def _build_coord_section(self, parent):
        sec = make_section(parent, "🌐 Chrome + 🖥️ 텔레그램 좌표")
        sec.pack(fill="both", expand=True, padx=4, pady=4)

        scroll = ScrollableFrame(sec, bg=THEME["card"])
        scroll.pack(fill="both", expand=True, padx=4, pady=4)

        inner = scroll.inner

        # 헤더
        hdr = tk.Frame(inner, bg=THEME["card"])
        hdr.pack(fill="x", padx=4, pady=(4, 2))
        for text, w in [("좌표 항목", 20), ("X", 7), ("Y", 7), ("", 8)]:
            tk.Label(hdr, text=text, bg=THEME["card"],
                     fg=THEME["subtext"],
                     font=FONT["small"], width=w,
                     anchor="w").pack(side="left", padx=2)

        ttk.Separator(inner, orient="horizontal").pack(fill="x", padx=4, pady=2)

        # 좌표 행
        for key_x, key_y, label, tooltip in COORD_FIELDS:
            row = tk.Frame(inner, bg=THEME["card"])
            row.pack(fill="x", padx=4, pady=3)

            # 레이블
            lbl = tk.Label(row, text=label,
                           bg=THEME["card"], fg=THEME["text"],
                           font=FONT["label"], width=20, anchor="w")
            lbl.pack(side="left", padx=(4, 6))

            # X 입력
            var_x = tk.StringVar(value="0")
            entry_x = tk.Entry(row, textvariable=var_x, width=7,
                               bg=THEME["input_bg"],
                               fg=THEME["text"],
                               insertbackground=THEME["text"],
                               font=FONT["small"], relief="flat")
            entry_x.pack(side="left", padx=2)

            # Y 입력
            var_y = tk.StringVar(value="0")
            entry_y = tk.Entry(row, textvariable=var_y, width=7,
                               bg=THEME["input_bg"],
                               fg=THEME["text"],
                               insertbackground=THEME["text"],
                               font=FONT["small"], relief="flat")
            entry_y.pack(side="left", padx=2)

            # Pick 버튼
            pick_btn = make_btn(
                row, "🎯 Pick",
                command=lambda kx=key_x, ky=key_y, vx=var_x, vy=var_y:
                    self._start_pick(kx, ky, vx, vy),
                color=THEME["border"],
                width=8
            )
            pick_btn.pack(side="left", padx=(4, 2))

            self._coord_vars[key_x] = var_x
            self._coord_vars[key_y] = var_y

        # 테스트 버튼
        test_frame = tk.Frame(inner, bg=THEME["card"])
        test_frame.pack(fill="x", padx=4, pady=8)

        tk.Label(test_frame, text="테스트 링크:",
                 bg=THEME["card"], fg=THEME["subtext"],
                 font=FONT["small"]).pack(side="left", padx=4)

        self._test_link_var = tk.StringVar(value="https://t.me/tonghab")
        tk.Entry(test_frame, textvariable=self._test_link_var,
                 width=30,
                 bg=THEME["input_bg"], fg=THEME["text"],
                 insertbackground=THEME["text"],
                 font=FONT["small"], relief="flat"
                 ).pack(side="left", padx=4)

        make_btn(test_frame, "🧪 전체 흐름 테스트",
                 self._run_test,
                 color=THEME["warning"]).pack(side="left", padx=4)

    def _build_delay_section(self, parent):
        sec = make_section(parent, "⏱️ 딜레이 설정 (초)")
        sec.pack(fill="x", padx=4, pady=4)

        inner = tk.Frame(sec, bg=THEME["card"])
        inner.pack(fill="x", padx=6, pady=4)

        for col_start in range(0, len(DELAY_FIELDS), 2):
            row_frame = tk.Frame(inner, bg=THEME["card"])
            row_frame.pack(fill="x", pady=2)

            for col_offset in range(2):
                idx = col_start + col_offset
                if idx >= len(DELAY_FIELDS):
                    break
                key, label, _ = DELAY_FIELDS[idx]

                cell = tk.Frame(row_frame, bg=THEME["card"])
                cell.pack(side="left", fill="x", expand=True, padx=4)

                tk.Label(cell, text=label,
                         bg=THEME["card"], fg=THEME["subtext"],
                         font=FONT["small"], anchor="w"
                         ).pack(side="left")

                var = tk.StringVar(value=str(DEFAULT_DELAYS.get(key, 1.0)))
                tk.Entry(cell, textvariable=var, width=6,
                         bg=THEME["input_bg"], fg=THEME["text"],
                         insertbackground=THEME["text"],
                         font=FONT["small"], relief="flat"
                         ).pack(side="left", padx=(4, 0))

                self._delay_vars[key] = var

    def _build_log_section(self, parent):
        sec = make_section(parent, "📋 로그")
        sec.pack(fill="both", expand=True, padx=4, pady=4)

        self._log = LogText(sec, height=8)
        self._log.pack(fill="both", expand=True, padx=4, pady=4)

        make_btn(sec, "🗑️ 지우기", self._log.clear,
                 color=THEME["border"]
                 ).pack(side="right", padx=6, pady=4)

    # ─── 이벤트 ──────────────────────────────────────────
    def _start_pick(self, key_x, key_y, var_x, var_y):
        """좌표 Pick 시작"""
        if self._picking:
            return
        self._picking = True
        self._log.append(fmt_log(f"🎯 [{key_x[:-2]}] 좌표 캡처 시작 (3초 후)"))

        def callback(result, mode):
            if mode == "countdown":
                self._log.append(fmt_log(f"  카운트다운: {result}초..."))
            elif mode == "done":
                x, y = result
                var_x.set(str(x))
                var_y.set(str(y))
                self._log.append(fmt_log(f"✅ 캡처 완료: ({x}, {y})"))
                self._picking = False
            elif mode == "error":
                self._log.append(fmt_log(f"❌ 캡처 오류: {result}"))
                self._picking = False

        pick_coordinate(callback)

    def _run_test(self):
        """테스트 링크로 전체 흐름 1회 실행"""
        link = self._test_link_var.get().strip()
        if not link:
            messagebox.showwarning("경고", "테스트 링크를 입력하세요.")
            return
        self._log.append(fmt_log(f"🧪 테스트 시작: {link}"))
        config = self.get_config()

        def _test():
            import time
            from utils.helpers import navigate_chrome
            import pyautogui
            try:
                self._log.append(fmt_log("① Chrome 주소창 클릭..."))
                navigate_chrome(link,
                                config["chrome_addr_x"],
                                config["chrome_addr_y"],
                                config["chrome_load_delay"])
                self._log.append(fmt_log("② Telegram 열기 버튼 클릭..."))
                pyautogui.click(config["chrome_open_btn_x"],
                                config["chrome_open_btn_y"])
                time.sleep(config["telegram_open_delay"])
                self._log.append(fmt_log("✅ 테스트 완료 (가입 클릭은 생략)"))
            except Exception as e:
                self._log.append(fmt_log(f"❌ 테스트 오류: {e}"))

        threading.Thread(target=_test, daemon=True).start()

    # ─── 설정 로드/저장 ──────────────────────────────────
    def get_config(self) -> dict:
        """현재 UI 값 → dict 반환"""
        conf = {}
        for key, var in self._coord_vars.items():
            conf[key] = safe_int(var.get(), 0)
        for key, var in self._delay_vars.items():
            conf[key] = safe_float(var.get(), 1.0)
        return conf

    def set_config(self, conf: dict):
        """dict → UI 값 설정"""
        for key, var in self._coord_vars.items():
            if key in conf:
                var.set(str(conf[key]))
        for key, var in self._delay_vars.items():
            if key in conf:
                var.set(str(conf[key]))

    def _load_config(self):
        """앱 시작 시 저장된 설정 불러오기"""
        path = get_config_path()
        if not os.path.exists(path):
            self._refresh_preset_list(["기본값"])
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            presets = data.get("presets", {})
            active = data.get("active_preset", "기본값")
            self._refresh_preset_list(list(presets.keys()))
            self._preset_var.set(active)
            if active in presets:
                self.set_config(presets[active])
        except Exception:
            pass

    def _save_config(self):
        """전체 설정 파일 저장"""
        path = get_config_path()
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
            else:
                data = {"presets": {}}

            active = self._preset_var.get()
            data["active_preset"] = active
            data["presets"][active] = self.get_config()

            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    def _save_preset(self):
        self._save_config()
        self._log.append(fmt_log(f"💾 프리셋 저장: {self._preset_var.get()}"))

    def _load_preset(self):
        path = get_config_path()
        if not os.path.exists(path):
            return
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            active = self._preset_var.get()
            presets = data.get("presets", {})
            if active in presets:
                self.set_config(presets[active])
                self._log.append(fmt_log(f"📂 프리셋 불러오기: {active}"))
        except Exception as e:
            messagebox.showerror("불러오기 오류", str(e))

    def _new_preset(self):
        name = simpledialog.askstring("새 프리셋", "프리셋 이름을 입력하세요:")
        if not name:
            return
        self._preset_var.set(name)
        current = list(self._preset_combo["values"])
        if name not in current:
            current.append(name)
            self._refresh_preset_list(current)
        self._save_config()
        self._log.append(fmt_log(f"➕ 새 프리셋 생성: {name}"))

    def _delete_preset(self):
        name = self._preset_var.get()
        if name == "기본값":
            messagebox.showwarning("경고", "기본값 프리셋은 삭제할 수 없습니다.")
            return
        if not messagebox.askyesno("삭제 확인", f"'{name}' 프리셋을 삭제할까요?"):
            return
        path = get_config_path()
        try:
            if os.path.exists(path):
                with open(path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                data["presets"].pop(name, None)
                data["active_preset"] = "기본값"
                with open(path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)

            presets = list(self._preset_combo["values"])
            if name in presets:
                presets.remove(name)
            self._refresh_preset_list(presets)
            self._preset_var.set("기본값")
            self._log.append(fmt_log(f"🗑️ 프리셋 삭제: {name}"))
        except Exception as e:
            messagebox.showerror("삭제 오류", str(e))

    def _on_preset_select(self, event=None):
        self._load_preset()

    def _refresh_preset_list(self, names: list):
        if "기본값" not in names:
            names.insert(0, "기본값")
        self._preset_combo["values"] = names


# ── D3: join_only_tab.py ─────────────────────────────
# [DUP_REMOVED] import os
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import tkinter as tk
# [DUP_REMOVED] from tkinter import ttk, filedialog, messagebox

# [MERGED] import config as cfg
# [DUP_CLEAN] from ui.widgets import LogText, make_btn, make_label, make_section
# [DUP_CLEAN] from utils.link_parser import parse_links, count_stats
# [DUP_CLEAN] from utils.excel_manager import ExcelManager, GroupRow
# [DUP_CLEAN] from utils.stop_manager import stop_controller
# [DUP_CLEAN] from utils.helpers import fmt_log, navigate_chrome, jitter_sleep, safe_sleep
# [DUP_CLEAN] from workers.join_worker import JoinWorker


class JoinOnlyTab(tk.Frame):
    """🎯 가입 전용 탭 — 합류만, 메시지 없음"""

    def __init__(self, parent, get_coord_config):
        super().__init__(parent, bg=THEME["bg"])
        self._get_coord_config = get_coord_config
        self._excel_manager = ExcelManager()
        self._worker = JoinWorker()
        self._worker_thread = None
        self._running = False
        self._parsed_links = []

        self._build_ui()

    # ─── UI ──────────────────────────────────────────────
    def _build_ui(self):
        # 좌우 분할
        paned = tk.Frame(self, bg=THEME["bg"])
        paned.pack(fill="both", expand=True, padx=8, pady=8)

        left = tk.Frame(paned, bg=THEME["bg"])
        left.pack(side="left", fill="both", expand=True, padx=(0, 4))

        right = tk.Frame(paned, bg=THEME["bg"], width=320)
        right.pack(side="right", fill="both")
        right.pack_propagate(False)

        self._build_link_section(left)
        self._build_control_section(right)

    def _build_link_section(self, parent):
        sec = make_section(parent, "📋 가입할 링크 목록")
        sec.pack(fill="both", expand=True)

        # 입력 방식 선택
        mode_row = tk.Frame(sec, bg=THEME["card"])
        mode_row.pack(fill="x", padx=10, pady=(8, 4))

        make_label(mode_row, "입력 방식:", bg=THEME["card"]).pack(
            side="left", padx=(0, 10))

        self._input_mode = tk.StringVar(value="text")
        for val, txt in [("text", "직접 입력"), ("excel", "Excel 파일")]:
            tk.Radiobutton(
                mode_row, text=txt,
                variable=self._input_mode, value=val,
                bg=THEME["card"],
                fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"],
                command=self._on_mode_change
            ).pack(side="left", padx=8)

        # 텍스트 입력
        self._text_frame = tk.Frame(sec, bg=THEME["card"])
        self._text_frame.pack(fill="both", expand=True, padx=10, pady=4)

        self._link_text = tk.Text(
            self._text_frame, height=22,
            bg=THEME["input_bg"], fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["mono"], relief="flat",
            wrap="none", padx=6, pady=4
        )
        sb_y = ttk.Scrollbar(self._text_frame, command=self._link_text.yview)
        sb_x = ttk.Scrollbar(self._text_frame, orient="horizontal",
                              command=self._link_text.xview)
        self._link_text.configure(yscrollcommand=sb_y.set,
                                  xscrollcommand=sb_x.set)
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        self._link_text.pack(fill="both", expand=True)

        # Excel 파일 선택 (숨김)
        self._excel_frame = tk.Frame(sec, bg=THEME["card"])
        self._excel_path_var = tk.StringVar(value="파일을 선택하세요...")
        tk.Label(
            self._excel_frame,
            textvariable=self._excel_path_var,
            bg=THEME["input_bg"],
            fg=THEME["subtext"],
            font=FONT["small"],
            anchor="w", relief="flat", padx=6
        ).pack(side="left", fill="x", expand=True, padx=(0, 4))
        make_btn(self._excel_frame, "📂 선택",
                 self._select_excel,
                 color=THEME["border"]).pack(side="right")

        # 버튼 바
        btn_bar = tk.Frame(sec, bg=THEME["card"])
        btn_bar.pack(fill="x", padx=10, pady=(4, 8))

        make_btn(btn_bar, "📋 파싱 & 중복제거",
                 self._parse_links,
                 color=THEME["border"]).pack(side="left", padx=4)

        self._stats_lbl = make_label(
            btn_bar, "",
            color=THEME["subtext"], bg=THEME["card"],
            font=FONT["small"]
        )
        self._stats_lbl.pack(side="left", padx=12)

        make_btn(btn_bar, "💾 Excel 내보내기",
                 self._export_excel,
                 color=THEME["accent"]).pack(side="right", padx=4)

    def _build_control_section(self, parent):
        # ── 옵션 ──
        opt_sec = make_section(parent, "⚙️ 옵션")
        opt_sec.pack(fill="x", padx=0, pady=(0, 6))

        opt_inner = tk.Frame(opt_sec, bg=THEME["card"])
        opt_inner.pack(fill="x", padx=10, pady=8)

        # 처리 대상
        make_label(opt_inner, "처리 대상:", bg=THEME["card"]).pack(
            anchor="w", pady=(0, 4))

        self._filter_var = tk.StringVar(value="미가입만")
        for val in ["미가입만", "전체", "실패만 재시도"]:
            tk.Radiobutton(
                opt_inner, text=val,
                variable=self._filter_var, value=val,
                bg=THEME["card"],
                fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"]
            ).pack(anchor="w", padx=8, pady=1)

        ttk.Separator(opt_inner, orient="horizontal").pack(
            fill="x", pady=8)

        # 딜레이 표시 (읽기 전용, CoordTab에서 가져옴)
        make_label(opt_inner,
                   "※ 딜레이는 좌표 설정 탭에서 조정",
                   color=THEME["subtext"],
                   font=FONT["small"],
                   bg=THEME["card"]).pack(anchor="w")

        # ── 실행 ──
        run_sec = make_section(parent, "🚀 실행")
        run_sec.pack(fill="both", expand=True, padx=0, pady=0)

        run_inner = tk.Frame(run_sec, bg=THEME["card"])
        run_inner.pack(fill="both", expand=True, padx=10, pady=8)

        # 시작/중단 버튼
        self._start_btn = make_btn(
            run_inner, "▶ 가입 시작  F8",
            self.start,
            color=THEME["success"]
        )
        self._start_btn.pack(fill="x", pady=(0, 4))

        self._stop_btn = make_btn(
            run_inner, "■ 중단  F9",
            self.stop,
            color=THEME["danger"]
        )
        self._stop_btn.pack(fill="x", pady=(0, 8))
        self._stop_btn.config(state="disabled")

        # 진행 바
        self._progress = ttk.Progressbar(
            run_inner, length=280, mode="determinate"
        )
        self._progress.pack(fill="x", pady=4)

        # 카운터 행
        cnt_row = tk.Frame(run_inner, bg=THEME["card"])
        cnt_row.pack(fill="x", pady=2)

        self._prog_lbl = make_label(
            cnt_row, "0 / 0", bg=THEME["card"])
        self._prog_lbl.pack(side="left")

        self._ok_lbl = make_label(
            cnt_row, "✅ 0",
            color=THEME["success"], bg=THEME["card"])
        self._ok_lbl.pack(side="left", padx=10)

        self._fail_lbl = make_label(
            cnt_row, "❌ 0",
            color=THEME["danger"], bg=THEME["card"])
        self._fail_lbl.pack(side="left")

        ttk.Separator(run_inner, orient="horizontal").pack(
            fill="x", pady=8)

        # 로그
        make_label(run_inner, "로그:", bg=THEME["card"]).pack(anchor="w")
        self._log = LogText(run_inner, height=16)
        self._log.pack(fill="both", expand=True, pady=4)

        make_btn(run_inner, "🗑️ 로그 지우기",
                 self._log.clear,
                 color=THEME["border"]).pack(side="right")

    # ─── 이벤트 ──────────────────────────────────────────
    def _on_mode_change(self):
        if self._input_mode.get() == "text":
            self._excel_frame.pack_forget()
            self._text_frame.pack(fill="both", expand=True, padx=10, pady=4)
        else:
            self._text_frame.pack_forget()
            self._excel_frame.pack(fill="x", padx=10, pady=4)

    def _parse_links(self):
        raw = self._link_text.get("1.0", "end")
        stats = count_stats(raw)
        self._parsed_links = parse_links(raw)
        self._stats_lbl.config(
            text=f"총 {stats['raw']}줄 → {stats['parsed']}개 (중복 {stats['duplicates']}개 제거)"
        )
        self._log.append(fmt_log(
            f"✅ 파싱 완료: {stats['parsed']}개 (중복 {stats['duplicates']}개 제거)"))

    def _select_excel(self):
        path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        self._excel_path_var.set(path)
        try:
            self._excel_manager.load(path)  # side-effect
            stats = self._excel_manager.get_stats()
            self._stats_lbl.config(
                text=f"전체:{stats['전체']}  미가입:{stats['미가입']}  "
                     f"가입완료:{stats['가입완료']}  실패:{stats['실패']}"
            )
            self._log.append(fmt_log(
                f"📂 {os.path.basename(path)} ({stats['전체']}개)"))
        except Exception as e:
            messagebox.showerror("로드 오류", str(e))

    def _export_excel(self):
        if not self._parsed_links:
            messagebox.showwarning("경고", "먼저 링크를 파싱하세요.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile="join_groups.xlsx"
        )
        if not path:
            return
        try:
            rows = self._excel_manager.create_from_links(
                self._parsed_links, path)
            self._log.append(fmt_log(
                f"💾 저장: {os.path.basename(path)} ({len(rows)}개)"))
            messagebox.showinfo("저장 완료",
                                f"{len(rows)}개 링크 저장\n경로: {path}")
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    # ─── 작업 제어 ────────────────────────────────────────
    def start(self):
        if self._running:
            return

        rows = self._get_rows()
        if not rows:
            messagebox.showwarning("경고",
                                   "처리할 링크가 없습니다.\n"
                                   "링크를 입력하거나 Excel을 불러오세요.")
            return

        coord_config = self._get_coord_config()
        # 메시지 발송 없음을 명시
        full_config = dict(coord_config)
        full_config["send_after_join"] = False

        self._running = True
        stop_controller.reset("join_only")
        cancel_event = stop_controller.get_event("join_only")

        self._start_btn.config(state="disabled")
        self._stop_btn.config(state="normal")
        self._log.append(fmt_log(f"▶ 가입 시작 — {len(rows)}개 링크"))

        def _run():
            self._worker.run(
                rows=rows,
                config=full_config,
                cancel_event=cancel_event,
                logger=lambda msg: self.after(
                    0, lambda m=msg: self._log.append(m)),
                excel_manager=self._excel_manager,
                progress_callback=lambda c, t, s, f: self.after(
                    0, lambda: self._update_progress(c, t, s, f)
                ),
            )
            self.after(0, self._on_done)

        self._worker_thread = threading.Thread(target=_run, daemon=True)
        self._worker_thread.start()

    def stop(self):
        stop_controller.stop("join_only")
        self._worker.stop()
        self._log.append(fmt_log("⏹ 중단 요청됨"))

    def _on_done(self):
        self._running = False
        self._start_btn.config(state="normal")
        self._stop_btn.config(state="disabled")

    def _update_progress(self, current, total, success, fail):
        self._prog_lbl.config(text=f"{current} / {total}")
        self._ok_lbl.config(text=f"✅ {success}")
        self._fail_lbl.config(text=f"❌ {fail}")
        if total > 0:
            self._progress["value"] = (current / total) * 100

    def _get_rows(self):
        if self._input_mode.get() == "text":
            if not self._parsed_links:
                self._parse_links()
            if not self._parsed_links:
                return []
            return [GroupRow(row_index=i + 2, link=lk)
                    for i, lk in enumerate(self._parsed_links)]
        else:
            fv = self._filter_var.get()
            if fv == "미가입만":
                return self._excel_manager.get_filtered("미가입만")
            elif fv == "실패만 재시도":
                return self._excel_manager.get_filtered("실패만")
            else:
                return self._excel_manager.get_filtered("전체")


# ── D4: join_tab.py ──────────────────────────────────
# [DUP_REMOVED] import os
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import tkinter as tk
# [DUP_REMOVED] from tkinter import ttk, filedialog, messagebox

# [MERGED] import config as cfg
# [DUP_CLEAN] from ui.widgets import LogText, make_btn, make_label, make_section
# [DUP_CLEAN] from utils.link_parser import parse_links, count_stats
# [DUP_CLEAN] from utils.excel_manager import ExcelManager
# [DUP_CLEAN] from utils.stop_manager import stop_controller
# [DUP_CLEAN] from utils.helpers import fmt_log
# [DUP_CLEAN] from workers.join_worker import JoinWorker


class JoinTab(tk.Frame):
    """🚪 자동 합류 탭"""

    def __init__(self, parent, get_coord_config):
        """
        get_coord_config: CoordTab.get_config 콜백
        """
        super().__init__(parent, bg=THEME["bg"])
        self._get_coord_config = get_coord_config
        self._excel_manager = ExcelManager()
        self._worker = JoinWorker()
        self._worker_thread = None
        self._running = False

        # 진행 상태
        self._total = 0
        self._current = 0
        self._success = 0
        self._fail = 0

        self._build_ui()

    # ─── UI 구성 ─────────────────────────────────────────
    def _build_ui(self):
        # 상단: 링크 입력
        self._build_input_section()
        # 중단: 옵션
        self._build_option_section()
        # 하단: 진행 + 로그
        self._build_progress_section()

    def _build_input_section(self):
        sec = make_section(self, "📋 링크 목록")
        sec.pack(fill="x", padx=8, pady=(8, 4))

        # 입력 방식 라디오
        mode_frame = tk.Frame(sec, bg=THEME["card"])
        mode_frame.pack(fill="x", padx=8, pady=(6, 2))

        self._input_mode = tk.StringVar(value="text")
        for val, txt in [("text", "직접 입력"), ("excel", "Excel 파일")]:
            tk.Radiobutton(
                mode_frame, text=txt,
                variable=self._input_mode, value=val,
                bg=THEME["card"],
                fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"],
                command=self._on_mode_change
            ).pack(side="left", padx=8)

        # 텍스트 입력 영역
        self._text_frame = tk.Frame(sec, bg=THEME["card"])
        self._text_frame.pack(fill="x", padx=8, pady=4)

        self._link_text = tk.Text(
            self._text_frame, height=8,
            bg=THEME["input_bg"], fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["mono"], relief="flat", wrap="none"
        )
        sb = ttk.Scrollbar(self._text_frame, command=self._link_text.yview)
        self._link_text.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        self._link_text.pack(fill="x")

        # Excel 파일 선택 영역
        self._excel_frame = tk.Frame(sec, bg=THEME["card"])

        self._excel_path_var = tk.StringVar(value="파일을 선택하세요...")
        tk.Label(self._excel_frame, textvariable=self._excel_path_var,
                 bg=THEME["input_bg"], fg=THEME["subtext"],
                 font=FONT["small"], anchor="w", relief="flat",
                 padx=6).pack(side="left", fill="x", expand=True, padx=(0, 4))
        make_btn(self._excel_frame, "📂 Excel 선택",
                 self._select_excel,
                 color=THEME["border"]).pack(side="right")

        # 버튼 바
        btn_bar = tk.Frame(sec, bg=THEME["card"])
        btn_bar.pack(fill="x", padx=8, pady=(2, 8))

        make_btn(btn_bar, "📋 파싱 & 중복제거",
                 self._parse_links,
                 color=THEME["border"]).pack(side="left", padx=4)

        self._stats_label = make_label(
            btn_bar, "링크: 0개",
            color=THEME["subtext"],
            bg=THEME["card"]
        )
        self._stats_label.pack(side="left", padx=12)

        make_btn(btn_bar, "💾 Excel로 내보내기",
                 self._export_excel,
                 color=THEME["accent"]).pack(side="right", padx=4)

    def _build_option_section(self):
        sec = make_section(self, "⚙️ 실행 옵션")
        sec.pack(fill="x", padx=8, pady=4)

        row1 = tk.Frame(sec, bg=THEME["card"])
        row1.pack(fill="x", padx=8, pady=4)

        # 처리 대상
        make_label(row1, "처리 대상:", bg=THEME["card"]).pack(
            side="left", padx=(0, 8))

        self._filter_var = tk.StringVar(value="미가입만")
        for val in ["미가입만", "전체", "실패만"]:
            tk.Radiobutton(
                row1, text=val,
                variable=self._filter_var, value=val,
                bg=THEME["card"],
                fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"]
            ).pack(side="left", padx=6)

        row2 = tk.Frame(sec, bg=THEME["card"])
        row2.pack(fill="x", padx=8, pady=(0, 4))

        # 통합 모드 체크박스
        self._combined_var = tk.BooleanVar(value=False)
        cb = tk.Checkbutton(
            row2, text="합류 후 즉시 메시지 발송 (통합 모드)",
            variable=self._combined_var,
            bg=THEME["card"],
            fg=THEME["warning"],
            activebackground=THEME["card"],
            activeforeground=THEME["warning"],
            selectcolor=THEME["input_bg"],
            highlightthickness=0,
            font=FONT["label"],
            command=self._toggle_combined
        )
        cb.pack(side="left", padx=(0, 12))

        # 통합 모드 메시지 입력
        self._combined_frame = tk.Frame(sec, bg=THEME["card"])
        self._combined_frame.pack(fill="x", padx=8, pady=(0, 6))

        tk.Label(self._combined_frame, text="발송 메시지:",
                 bg=THEME["card"], fg=THEME["subtext"],
                 font=FONT["small"]).pack(side="left", padx=(0, 4))

        self._join_msg_var = tk.StringVar()
        tk.Entry(
            self._combined_frame,
            textvariable=self._join_msg_var,
            width=50,
            bg=THEME["input_bg"], fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["small"], relief="flat"
        ).pack(side="left", padx=4)

        # 이미지 옵션
        img_frame = tk.Frame(self._combined_frame, bg=THEME["card"])
        img_frame.pack(side="left", padx=8)

        tk.Label(img_frame, text="이미지:",
                 bg=THEME["card"], fg=THEME["subtext"],
                 font=FONT["small"]).pack(side="left")

        self._img_mode_var = tk.StringVar(value="none")
        for val, txt in [("none", "없음"), ("files", "첨부")]:
            tk.Radiobutton(
                img_frame, text=txt,
                variable=self._img_mode_var, value=val,
                bg=THEME["card"],
                fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["small"]
            ).pack(side="left", padx=4)

        self._combined_frame.pack_forget()  # 초기 숨김

    def _build_progress_section(self):
        sec = make_section(self, "🚀 실행")
        sec.pack(fill="both", expand=True, padx=8, pady=(4, 8))

        # 버튼 바
        btn_bar = tk.Frame(sec, bg=THEME["card"])
        btn_bar.pack(fill="x", padx=8, pady=6)

        self._start_btn = make_btn(
            btn_bar, "▶ 합류 시작  F8",
            self.start,
            color=THEME["success"], width=18
        )
        self._start_btn.pack(side="left", padx=4)

        self._stop_btn = make_btn(
            btn_bar, "■ 중단  F9",
            self.stop,
            color=THEME["danger"], width=12
        )
        self._stop_btn.pack(side="left", padx=4)
        self._stop_btn.config(state="disabled")

        # 진행 표시
        prog_frame = tk.Frame(sec, bg=THEME["card"])
        prog_frame.pack(fill="x", padx=8, pady=4)

        self._progress = ttk.Progressbar(
            prog_frame, length=400, mode="determinate"
        )
        self._progress.pack(side="left", padx=(0, 8), fill="x", expand=True)

        self._prog_label = make_label(
            prog_frame, "0 / 0",
            color=THEME["text"],
            bg=THEME["card"]
        )
        self._prog_label.pack(side="left", padx=4)

        self._success_label = make_label(
            prog_frame, "✅ 0",
            color=THEME["success"],
            bg=THEME["card"]
        )
        self._success_label.pack(side="left", padx=8)

        self._fail_label = make_label(
            prog_frame, "❌ 0",
            color=THEME["danger"],
            bg=THEME["card"]
        )
        self._fail_label.pack(side="left", padx=4)

        # 로그
        self._log = LogText(sec, height=12)
        self._log.pack(fill="both", expand=True, padx=8, pady=(4, 4))

        log_btn_bar = tk.Frame(sec, bg=THEME["card"])
        log_btn_bar.pack(fill="x", padx=8, pady=(0, 6))
        make_btn(log_btn_bar, "🗑️ 로그 지우기",
                 self._log.clear,
                 color=THEME["border"]).pack(side="right", padx=4)

    # ─── 이벤트 ──────────────────────────────────────────
    def _on_mode_change(self):
        mode = self._input_mode.get()
        if mode == "text":
            self._excel_frame.pack_forget()
            self._text_frame.pack(fill="x", padx=8, pady=4)
        else:
            self._text_frame.pack_forget()
            self._excel_frame.pack(fill="x", padx=8, pady=4)

    def _toggle_combined(self):
        if self._combined_var.get():
            self._combined_frame.pack(fill="x", padx=8, pady=(0, 6))
        else:
            self._combined_frame.pack_forget()

    def _parse_links(self):
        raw = self._link_text.get("1.0", "end")
        stats = count_stats(raw)
        links = parse_links(raw)

        self._parsed_links = links
        self._stats_label.config(
            text=f"총 {stats['raw']}줄 → 유효 {stats['parsed']}개 (중복제거 {stats['duplicates']}개)"
        )
        self._log.append(fmt_log(
            f"✅ 파싱 완료: {stats['parsed']}개 링크 (중복 {stats['duplicates']}개 제거)"
        ))

    def _select_excel(self):
        path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        self._excel_path_var.set(path)
        try:
            self._excel_manager.load(path)  # side-effect
            stats = self._excel_manager.get_stats()
            self._stats_label.config(
                text=f"전체: {stats['전체']}개  "
                     f"미가입: {stats['미가입']}  "
                     f"가입완료: {stats['가입완료']}  "
                     f"실패: {stats['실패']}"
            )
            self._log.append(fmt_log(f"📂 Excel 로드: {os.path.basename(path)} ({stats['전체']}개)"))
        except Exception as e:
            messagebox.showerror("로드 오류", str(e))

    def _export_excel(self):
        links = getattr(self, "_parsed_links", [])
        if not links:
            messagebox.showwarning("경고", "먼저 링크를 파싱하세요.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile="telegram_groups.xlsx"
        )
        if not path:
            return
        try:
            rows = self._excel_manager.create_from_links(links, path)
            self._log.append(fmt_log(f"💾 Excel 저장: {os.path.basename(path)} ({len(rows)}개)"))
            messagebox.showinfo("저장 완료", f"{len(rows)}개 링크가 저장되었습니다.")
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    # ─── 작업 제어 ────────────────────────────────────────
    def start(self):
        if self._running:
            return

        # 링크/Excel 준비
        rows = self._get_rows()
        if not rows:
            messagebox.showwarning("경고", "처리할 링크가 없습니다.\n링크를 입력하거나 Excel을 불러오세요.")
            return

        # 좌표 설정 가져오기
        coord_config = self._get_coord_config()

        # 통합 모드 설정 병합
        full_config = dict(coord_config)
        if self._combined_var.get():
            full_config["send_after_join"] = True
            full_config["join_msg"] = self._join_msg_var.get()
            full_config["image_mode"] = self._img_mode_var.get()

        # 상태 초기화
        self._running = True
        self._total = len(rows)
        self._current = 0
        self._success = 0
        self._fail = 0
        self._update_progress(0, self._total, 0, 0)

        stop_controller.reset("join")
        cancel_event = stop_controller.get_event("join")

        self._start_btn.config(state="disabled")
        self._stop_btn.config(state="normal")
        self._log.append(fmt_log(f"▶ 합류 시작 — {self._total}개 링크"))

        def _run():
            self._worker.run(
                rows=rows,
                config=full_config,
                cancel_event=cancel_event,
                logger=lambda msg: self.after(0, lambda m=msg: self._log.append(m)),
                excel_manager=self._excel_manager,
                progress_callback=lambda c, t, s, f: self.after(
                    0, lambda: self._update_progress(c, t, s, f)
                ),
            )
            self.after(0, self._on_complete)

        self._worker_thread = threading.Thread(target=_run, daemon=True)
        self._worker_thread.start()

    def stop(self):
        stop_controller.stop("join")
        self._worker.stop()
        self._log.append(fmt_log("⏹ 중단 요청됨"))

    def _on_complete(self):
        self._running = False
        self._start_btn.config(state="normal")
        self._stop_btn.config(state="disabled")

    def _update_progress(self, current, total, success, fail):
        self._prog_label.config(text=f"{current} / {total}")
        self._success_label.config(text=f"✅ {success}")
        self._fail_label.config(text=f"❌ {fail}")
        if total > 0:
            self._progress["value"] = (current / total) * 100

    def _get_rows(self):
        """현재 입력 모드에 따라 처리할 행 목록 반환"""
        filter_val = self._filter_var.get()

        if self._input_mode.get() == "text":
            # 텍스트 입력 → 링크 파싱 후 임시 GroupRow 생성
            if not hasattr(self, "_parsed_links") or not self._parsed_links:
                self._parse_links()
            links = getattr(self, "_parsed_links", [])
            if not links:
                return []
            from utils.excel_manager import GroupRow
            return [GroupRow(row_index=i + 2, link=lk)
                    for i, lk in enumerate(links)]
        else:
            # Excel 파일
            if not self._excel_manager.rows:
                return []
            return self._excel_manager.get_filtered(filter_val)


# ── D5: message_tab.py ───────────────────────────────
# [DUP_REMOVED] import os
# [DUP_REMOVED] import threading
# [DUP_REMOVED] import tkinter as tk
# [DUP_REMOVED] from tkinter import ttk, filedialog, messagebox

# [MERGED] import config as cfg
# [DUP_CLEAN] from ui.widgets import LogText, make_btn, make_label, make_section
# [DUP_CLEAN] from utils.excel_manager import ExcelManager, GroupRow
# [DUP_CLEAN] from utils.stop_manager import stop_controller
# [DUP_CLEAN] from utils.helpers import fmt_log
# [DUP_CLEAN] from workers.message_worker import MessageWorker


class MessageTab(tk.Frame):
    """💬 자동 메시지 발송 탭"""

    def __init__(self, parent, get_coord_config):
        super().__init__(parent, bg=THEME["bg"])
        self._get_coord_config = get_coord_config
        self._excel_manager = ExcelManager()
        self._worker = MessageWorker()
        self._worker_thread = None
        self._running = False
        self._rows = []

        self._build_ui()

    # ─── UI 구성 ─────────────────────────────────────────
    def _build_ui(self):
        # 좌우 분할 (그룹 목록 | 설정)
        main = tk.Frame(self, bg=THEME["bg"])
        main.pack(fill="both", expand=True)

        # 왼쪽: 그룹 목록
        left = tk.Frame(main, bg=THEME["bg"])
        left.pack(side="left", fill="both", expand=True, padx=(8, 4), pady=8)

        # 오른쪽: 설정 + 진행
        right = tk.Frame(main, bg=THEME["bg"], width=380)
        right.pack(side="right", fill="both", padx=(4, 8), pady=8)
        right.pack_propagate(False)

        self._build_list_section(left)
        self._build_setting_section(right)
        self._build_run_section(right)

    def _build_list_section(self, parent):
        sec = make_section(parent, "📊 그룹 목록")
        sec.pack(fill="both", expand=True)

        # 상단 버튼
        btn_bar = tk.Frame(sec, bg=THEME["card"])
        btn_bar.pack(fill="x", padx=8, pady=6)

        make_btn(btn_bar, "📂 Excel 불러오기",
                 self._load_excel,
                 color=THEME["accent"]).pack(side="left", padx=4)
        make_btn(btn_bar, "💾 Excel 저장",
                 self._save_excel,
                 color=THEME["border"]).pack(side="left", padx=4)

        self._list_stats = make_label(
            btn_bar, "",
            color=THEME["subtext"], bg=THEME["card"]
        )
        self._list_stats.pack(side="left", padx=12)

        # 처리 대상 필터
        filter_frame = tk.Frame(sec, bg=THEME["card"])
        filter_frame.pack(fill="x", padx=8, pady=(0, 4))

        make_label(filter_frame, "처리 대상:", bg=THEME["card"]).pack(
            side="left", padx=(0, 8))

        self._filter_var = tk.StringVar(value="가입완료만")
        for val in ["가입완료만", "전체", "재발송 포함"]:
            tk.Radiobutton(
                filter_frame, text=val,
                variable=self._filter_var, value=val,
                bg=THEME["card"],
                fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"]
            ).pack(side="left", padx=6)

        # Treeview
        tree_frame = tk.Frame(sec, bg=THEME["card"])
        tree_frame.pack(fill="both", expand=True, padx=8, pady=4)

        cols = ("링크", "상태", "커스텀메시지", "이미지")
        self._tree = ttk.Treeview(
            tree_frame, columns=cols,
            show="headings", height=12
        )

        style = ttk.Style()
        style.configure("Treeview",
                        background=THEME["input_bg"],
                        foreground=THEME["text"],
                        fieldbackground=THEME["input_bg"],
                        rowheight=22)
        style.configure("Treeview.Heading",
                        background=THEME["border"],
                        foreground=THEME["text"])

        widths = [280, 80, 180, 80]
        for col, w in zip(cols, widths):
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, anchor="w")

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        self._tree.pack(fill="both", expand=True)

    def _build_setting_section(self, parent):
        # 기본 메시지
        sec = make_section(parent, "✉️ 메시지 설정")
        sec.pack(fill="x", padx=0, pady=(0, 4))

        inner = tk.Frame(sec, bg=THEME["card"])
        inner.pack(fill="x", padx=8, pady=6)

        make_label(inner, "기본 메시지:", bg=THEME["card"]).pack(
            anchor="w")
        self._default_msg = tk.Text(
            inner, height=4,
            bg=THEME["input_bg"], fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["label"], relief="flat"
        )
        self._default_msg.pack(fill="x", pady=(2, 4))

        hint = "플레이스홀더: {{custom_msg}}  {{date}}  {{link}}"
        make_label(inner, hint,
                   color=THEME["subtext"],
                   font=FONT["small"],
                   bg=THEME["card"]).pack(anchor="w")

        # 전송 방식
        send_frame = tk.Frame(inner, bg=THEME["card"])
        send_frame.pack(fill="x", pady=(6, 0))
        make_label(send_frame, "전송 방식:", bg=THEME["card"]).pack(
            side="left", padx=(0, 8))
        self._send_method = tk.StringVar(value="enter")
        for val, txt in [("enter", "Enter 키"), ("click", "전송버튼 클릭")]:
            tk.Radiobutton(
                send_frame, text=txt,
                variable=self._send_method, value=val,
                bg=THEME["card"], fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"]
            ).pack(side="left", padx=6)

        # 이미지 설정
        img_sec = make_section(parent, "🖼️ 이미지/움짤 설정")
        img_sec.pack(fill="x", padx=0, pady=4)

        img_inner = tk.Frame(img_sec, bg=THEME["card"])
        img_inner.pack(fill="x", padx=8, pady=6)

        # 방식
        mode_frame = tk.Frame(img_inner, bg=THEME["card"])
        mode_frame.pack(fill="x", pady=2)
        make_label(mode_frame, "방식:", bg=THEME["card"]).pack(
            side="left", padx=(0, 8))
        self._img_mode = tk.StringVar(value="none")
        for val, txt in [("none", "없음"), ("files", "파일 다이얼로그")]:
            tk.Radiobutton(
                mode_frame, text=txt,
                variable=self._img_mode, value=val,
                bg=THEME["card"], fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"]
            ).pack(side="left", padx=6)

        # 타이밍
        timing_frame = tk.Frame(img_inner, bg=THEME["card"])
        timing_frame.pack(fill="x", pady=2)
        make_label(timing_frame, "타이밍:", bg=THEME["card"]).pack(
            side="left", padx=(0, 8))
        self._img_timing = tk.StringVar(value="after")
        for val, txt in [("before", "메시지 전"), ("after", "메시지 후"),
                         ("only", "이미지만")]:
            tk.Radiobutton(
                timing_frame, text=txt,
                variable=self._img_timing, value=val,
                bg=THEME["card"], fg=THEME["text"],
                activebackground=THEME["border"],
                activeforeground="#ffffff",
                selectcolor=THEME["accent"],
                highlightthickness=0,
                font=FONT["label"]
            ).pack(side="left", padx=6)

        # 이미지 폴더
        folder_frame = tk.Frame(img_inner, bg=THEME["card"])
        folder_frame.pack(fill="x", pady=4)
        make_label(folder_frame, "폴더:", bg=THEME["card"]).pack(
            side="left", padx=(0, 4))
        self._img_folder_var = tk.StringVar(value="")
        tk.Entry(
            folder_frame, textvariable=self._img_folder_var,
            width=22,
            bg=THEME["input_bg"], fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["small"], relief="flat"
        ).pack(side="left", padx=4)
        make_btn(folder_frame, "📂",
                 self._select_img_folder,
                 color=THEME["border"], width=3
                 ).pack(side="left")

        note = make_label(img_inner,
                          "※ Excel D열 개별 경로 우선 적용",
                          color=THEME["subtext"],
                          font=FONT["small"],
                          bg=THEME["card"])
        note.pack(anchor="w", pady=(2, 0))

    def _build_run_section(self, parent):
        sec = make_section(parent, "🚀 실행")
        sec.pack(fill="both", expand=True, padx=0, pady=4)

        btn_bar = tk.Frame(sec, bg=THEME["card"])
        btn_bar.pack(fill="x", padx=8, pady=6)

        self._start_btn = make_btn(
            btn_bar, "▶ 발송 시작  F8",
            self.start,
            color=THEME["success"], width=16
        )
        self._start_btn.pack(side="left", padx=4)

        self._stop_btn = make_btn(
            btn_bar, "■ 중단  F9",
            self.stop,
            color=THEME["danger"], width=10
        )
        self._stop_btn.pack(side="left", padx=4)
        self._stop_btn.config(state="disabled")

        # 진행
        prog_frame = tk.Frame(sec, bg=THEME["card"])
        prog_frame.pack(fill="x", padx=8, pady=4)

        self._progress = ttk.Progressbar(
            prog_frame, length=300, mode="determinate"
        )
        self._progress.pack(fill="x", expand=True, pady=2)

        stat_row = tk.Frame(sec, bg=THEME["card"])
        stat_row.pack(fill="x", padx=8)

        self._prog_label = make_label(
            stat_row, "0 / 0", bg=THEME["card"])
        self._prog_label.pack(side="left")
        self._success_label = make_label(
            stat_row, "✅ 0",
            color=THEME["success"], bg=THEME["card"])
        self._success_label.pack(side="left", padx=12)
        self._fail_label = make_label(
            stat_row, "❌ 0",
            color=THEME["danger"], bg=THEME["card"])
        self._fail_label.pack(side="left")

        # 로그
        self._log = LogText(sec, height=10)
        self._log.pack(fill="both", expand=True, padx=8, pady=4)

        make_btn(sec, "🗑️ 로그 지우기", self._log.clear,
                 color=THEME["border"]
                 ).pack(side="right", padx=8, pady=4)

    # ─── 이벤트 ──────────────────────────────────────────
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        try:
            self._rows = self._excel_manager.load(path)
            self._refresh_tree()
            stats = self._excel_manager.get_stats()
            self._list_stats.config(
                text=f"전체:{stats['전체']}  "
                     f"가입완료:{stats['가입완료']}  "
                     f"발송완료:{stats['발송완료']}"
            )
            self._log.append(fmt_log(
                f"📂 {os.path.basename(path)} 로드 ({stats['전체']}개)"))
        except Exception as e:
            messagebox.showerror("로드 오류", str(e))

    def _save_excel(self):
        if not self._excel_manager.filepath:
            path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel 파일", "*.xlsx")]
            )
            if not path:
                return
            self._excel_manager.filepath = path
        try:
            self._excel_manager.save_all()
            self._log.append(fmt_log("💾 Excel 저장 완료"))
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    def _select_img_folder(self):
        path = filedialog.askdirectory(title="이미지 폴더 선택")
        if path:
            self._img_folder_var.set(path)

    def _refresh_tree(self):
        self._tree.delete(*self._tree.get_children())
        for row in self._rows:
            has_img = "✅" if row.image_path else "❌"
            self._tree.insert("", "end", values=(
                row.link, row.status,
                row.custom_msg or "(기본값)", has_img
            ))

    # ─── 작업 제어 ────────────────────────────────────────
    def start(self):
        if self._running:
            return

        rows = self._excel_manager.get_filtered(self._filter_var.get())
        if not rows:
            messagebox.showwarning("경고",
                                   "처리할 그룹이 없습니다.\nExcel 파일을 불러오세요.")
            return

        coord_config = self._get_coord_config()
        full_config = dict(coord_config)
        full_config.update({
            "default_msg": self._default_msg.get("1.0", "end").strip(),
            "send_method": self._send_method.get(),
            "image_mode": self._img_mode.get(),
            "image_timing": self._img_timing.get(),
            "image_folder": self._img_folder_var.get(),
        })

        self._running = True
        stop_controller.reset("message")
        cancel_event = stop_controller.get_event("message")

        self._start_btn.config(state="disabled")
        self._stop_btn.config(state="normal")
        self._log.append(fmt_log(f"▶ 발송 시작 — {len(rows)}개"))

        def _run():
            self._worker.run(
                rows=rows,
                config=full_config,
                cancel_event=cancel_event,
                logger=lambda msg: self.after(
                    0, lambda m=msg: self._log.append(m)),
                excel_manager=self._excel_manager,
                progress_callback=lambda c, t, s, f: self.after(
                    0, lambda: self._update_progress(c, t, s, f)
                ),
            )
            self.after(0, self._on_complete)

        self._worker_thread = threading.Thread(target=_run, daemon=True)
        self._worker_thread.start()

    def stop(self):
        stop_controller.stop("message")
        self._worker.stop()
        self._log.append(fmt_log("⏹ 중단 요청됨"))

    def _on_complete(self):
        self._running = False
        self._start_btn.config(state="normal")
        self._stop_btn.config(state="disabled")
        self._refresh_tree()

    def _update_progress(self, current, total, success, fail):
        self._prog_label.config(text=f"{current} / {total}")
        self._success_label.config(text=f"✅ {success}")
        self._fail_label.config(text=f"❌ {fail}")
        if total > 0:
            self._progress["value"] = (current / total) * 100


# ── D6: manager_tab.py ───────────────────────────────
# [DUP_REMOVED] import os
# [DUP_REMOVED] import tkinter as tk
# [DUP_REMOVED] from tkinter import ttk, filedialog, messagebox

# [MERGED] import config as cfg
# [DUP_CLEAN] from ui.widgets import LogText, make_btn, make_label, make_section
# [DUP_CLEAN] from utils.link_parser import parse_links, count_stats
# [DUP_CLEAN] from utils.excel_manager import ExcelManager, GroupRow
# [DUP_CLEAN] from utils.helpers import fmt_log


class ManagerTab(tk.Frame):
    """📊 그룹 통합 관리 탭"""

    def __init__(self, parent):
        super().__init__(parent, bg=THEME["bg"])
        self._excel_manager = ExcelManager()
        self._rows = []
        self._filtered_rows = []
        self._build_ui()

    # ─── UI 구성 ─────────────────────────────────────────
    def _build_ui(self):
        self._build_stats_bar()
        self._build_main_area()
        self._build_add_section()

    def _build_stats_bar(self):
        bar = tk.Frame(self, bg=THEME["card"])
        bar.pack(fill="x", padx=8, pady=(8, 4))

        self._stat_labels = {}
        items = [
            ("전체",    THEME["text"]),
            ("미가입",  THEME["subtext"]),
            ("가입완료", THEME["warning"]),
            ("발송완료", THEME["success"]),
            ("실패",    THEME["danger"]),
            ("건너뜀",  "#888888"),
        ]
        for key, color in items:
            cell = tk.Frame(bar, bg=THEME["card"])
            cell.pack(side="left", padx=16, pady=8)
            tk.Label(cell, text=key,
                     bg=THEME["card"],
                     fg=THEME["subtext"],
                     font=FONT["small"]).pack()
            lbl = tk.Label(cell, text="0",
                           bg=THEME["card"],
                           fg=color,
                           font=("맑은 고딕", 16, "bold"))
            lbl.pack()
            self._stat_labels[key] = lbl

    def _build_main_area(self):
        main = tk.Frame(self, bg=THEME["bg"])
        main.pack(fill="both", expand=True, padx=8, pady=4)

        left = tk.Frame(main, bg=THEME["bg"])
        left.pack(side="left", fill="both", expand=True, padx=(0, 4))

        right = tk.Frame(main, bg=THEME["bg"], width=260)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        self._build_list_area(left)
        self._build_action_panel(right)

    def _build_list_area(self, parent):
        sec = make_section(parent, "📋 그룹 목록")
        sec.pack(fill="both", expand=True)

        # ── 현재 파일 경로 표시 ──────────────────────────
        path_bar = tk.Frame(sec, bg=THEME["input_bg"])
        path_bar.pack(fill="x", padx=8, pady=(6, 2))

        tk.Label(path_bar, text="📄 현재 파일:",
                 bg=THEME["input_bg"],
                 fg=THEME["subtext"],
                 font=FONT["small"],
                 padx=6).pack(side="left")

        self._filepath_lbl = tk.Label(
            path_bar,
            text="(파일 없음)",
            bg=THEME["input_bg"],
            fg=THEME["warning"],
            font=FONT["small"],
            anchor="w"
        )
        self._filepath_lbl.pack(side="left", fill="x", expand=True)

        # ── 필터 + 검색 바 ──────────────────────────────
        filter_bar = tk.Frame(sec, bg=THEME["card"])
        filter_bar.pack(fill="x", padx=8, pady=4)

        # ▶ 상태 필터 레이블
        tk.Label(
            filter_bar, text="상태 필터:",
            bg=THEME["card"],
            fg=THEME["text"],          # ← 흰색
            font=FONT["label"]
        ).pack(side="left", padx=(4, 6))

        # ▶ 상태 필터 OptionMenu (ttk.Combobox 대신 — 색상 완전 제어)
        self._filter_var = tk.StringVar(value="전체")
        filter_opts = ["전체", "미가입", "가입완료", "발송완료", "실패", "건너뜀"]
        filter_menu = tk.OptionMenu(
            filter_bar, self._filter_var, *filter_opts,
            command=self._apply_filter
        )
        filter_menu.config(
            bg=THEME["border"],
            fg=THEME["text"],           # ← 흰색
            activebackground=THEME["accent"],
            activeforeground="#ffffff",
            highlightthickness=0,
            relief="flat",
            font=FONT["label"],
            width=8,
        )
        filter_menu["menu"].config(
            bg=THEME["card"],
            fg=THEME["text"],           # ← 흰색
            activebackground=THEME["accent"],
            activeforeground="#ffffff",
            font=FONT["label"],
        )
        filter_menu.pack(side="left", padx=4)

        # ▶ 검색 레이블
        tk.Label(
            filter_bar, text="🔍",
            bg=THEME["card"],
            fg=THEME["text"],
            font=FONT["label"]
        ).pack(side="left", padx=(12, 4))

        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", self._apply_filter)
        tk.Entry(
            filter_bar,
            textvariable=self._search_var,
            width=26,
            bg=THEME["input_bg"],
            fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["small"],
            relief="flat"
        ).pack(side="left", padx=4)

        make_btn(
            filter_bar, "🔄 새로고침",
            self._refresh,
            color=THEME["border"]
        ).pack(side="right", padx=4)

        # ── Treeview ────────────────────────────────────
        tree_frame = tk.Frame(sec, bg=THEME["card"])
        tree_frame.pack(fill="both", expand=True, padx=8, pady=4)

        cols = ("#", "링크", "상태", "처리일시", "비고")
        self._tree = ttk.Treeview(
            tree_frame, columns=cols,
            show="headings", height=18,
            selectmode="extended"
        )

        widths  = [40, 300, 90, 150, 140]
        anchors = ["center", "w", "center", "center", "w"]
        for col, w, a in zip(cols, widths, anchors):
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, anchor=a)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical",
                            command=self._tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal",
                            command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set,
                             xscrollcommand=hsb.set)
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")
        self._tree.pack(fill="both", expand=True)

        self._tree.bind("<Double-Button-1>", self._on_double_click)

        # 행 색상 태그
        self._tree.tag_configure("joined",
            background="#1a2a10", foreground=THEME["warning"])
        self._tree.tag_configure("sent",
            background="#0d2a1a", foreground=THEME["success"])
        self._tree.tag_configure("failed",
            background="#2a0d0d", foreground=THEME["danger"])
        self._tree.tag_configure("skipped",
            background=THEME["input_bg"], foreground="#888888")
        self._tree.tag_configure("default",
            background=THEME["input_bg"], foreground=THEME["text"])

    def _build_action_panel(self, parent):
        sec = make_section(parent, "🛠️ 액션")
        sec.pack(fill="both", expand=True)

        inner = tk.Frame(sec, bg=THEME["card"])
        inner.pack(fill="both", expand=True, padx=8, pady=6)

        # ── Excel I/O ────────────────────────────────────
        tk.Label(inner, text="── Excel ──",
                 bg=THEME["card"],
                 fg=THEME["subtext"],
                 font=FONT["small"]).pack(pady=(4, 2))

        make_btn(inner, "📥 Excel 불러오기",
                 self._load_excel,
                 color=THEME["accent"],
                 width=20).pack(pady=3, fill="x")

        make_btn(inner, "📤 다른 이름으로 저장",
                 self._save_excel_as,
                 color=THEME["success"],
                 width=20).pack(pady=3, fill="x")

        make_btn(inner, "💾 현재 파일에 저장",
                 self._save_excel_current,
                 color=THEME["border"],
                 width=20).pack(pady=3, fill="x")

        ttk.Separator(inner, orient="horizontal").pack(fill="x", pady=8)

        # ── 선택 항목 ────────────────────────────────────
        tk.Label(inner, text="── 선택 항목 ──",
                 bg=THEME["card"],
                 fg=THEME["subtext"],
                 font=FONT["small"]).pack(pady=(0, 2))

        # 상태 변경 OptionMenu (색상 완전 제어)
        tk.Label(inner, text="변경할 상태:",
                 bg=THEME["card"],
                 fg=THEME["text"],
                 font=FONT["label"]).pack(anchor="w", pady=(4, 0))

        self._new_status_var = tk.StringVar(value="미가입")
        status_opts = ["미가입", "가입완료", "발송완료", "실패", "건너뜀"]
        status_menu = tk.OptionMenu(
            inner, self._new_status_var, *status_opts
        )
        status_menu.config(
            bg=THEME["border"],
            fg=THEME["text"],
            activebackground=THEME["accent"],
            activeforeground="#ffffff",
            highlightthickness=0,
            relief="flat",
            font=FONT["label"],
            width=12,
        )
        status_menu["menu"].config(
            bg=THEME["card"],
            fg=THEME["text"],
            activebackground=THEME["accent"],
            activeforeground="#ffffff",
            font=FONT["label"],
        )
        status_menu.pack(pady=4, fill="x")

        make_btn(inner, "✏️ 상태 일괄 변경",
                 self._bulk_change_status,
                 color=THEME["warning"],
                 width=20).pack(pady=3, fill="x")

        make_btn(inner, "🗑️ 선택 항목 삭제",
                 self._delete_selected,
                 color=THEME["danger"],
                 width=20).pack(pady=3, fill="x")

        ttk.Separator(inner, orient="horizontal").pack(fill="x", pady=8)

        # ── 기타 ────────────────────────────────────────
        tk.Label(inner, text="── 기타 ──",
                 bg=THEME["card"],
                 fg=THEME["subtext"],
                 font=FONT["small"]).pack(pady=(0, 2))

        make_btn(inner, "🔗 중복 링크 제거",
                 self._remove_duplicates,
                 color=THEME["border"],
                 width=20).pack(pady=3, fill="x")

        make_btn(inner, "📋 링크 전체 복사",
                 self._copy_all_links,
                 color=THEME["border"],
                 width=20).pack(pady=3, fill="x")

        # 로그
        self._log = LogText(inner, height=7)
        self._log.pack(fill="both", expand=True, pady=(8, 0))

    def _build_add_section(self):
        sec = make_section(self, "➕ 링크 추가")
        sec.pack(fill="x", padx=8, pady=(0, 8))

        inner = tk.Frame(sec, bg=THEME["card"])
        inner.pack(fill="x", padx=8, pady=6)

        tk.Label(inner, text="링크 (한 줄에 하나씩):",
                 bg=THEME["card"],
                 fg=THEME["text"],
                 font=FONT["label"]).pack(anchor="w")

        input_row = tk.Frame(inner, bg=THEME["card"])
        input_row.pack(fill="x", pady=4)

        self._add_text = tk.Text(
            input_row, height=3,
            bg=THEME["input_bg"],
            fg=THEME["text"],
            insertbackground=THEME["text"],
            font=FONT["mono"], relief="flat", wrap="none"
        )
        self._add_text.pack(side="left", fill="x", expand=True)

        make_btn(inner, "➕ 추가 & 파싱",
                 self._add_links,
                 color=THEME["success"],
                 width=16).pack(side="right", padx=(8, 0), pady=4)

    # ─── 파일 경로 업데이트 ──────────────────────────────
    def _update_filepath_label(self, path: str = ""):
        if path:
            display = path if len(path) <= 70 else "..." + path[-67:]
            self._filepath_lbl.config(
                text=f"✅ {display}",
                fg=THEME["success"]
            )
        else:
            self._filepath_lbl.config(
                text="(파일 없음 — 저장 시 다이얼로그 표시)",
                fg=THEME["warning"]
            )

    # ─── 이벤트 ──────────────────────────────────────────
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="Excel 파일 선택",
            filetypes=[("Excel 파일", "*.xlsx"), ("모든 파일", "*.*")]
        )
        if not path:
            return
        try:
            self._rows = self._excel_manager.load(path)
            self._apply_filter()
            self._update_stats()
            self._update_filepath_label(path)
            self._log.append(fmt_log(
                f"📥 불러오기: {os.path.basename(path)} ({len(self._rows)}개)"))
        except Exception as e:
            messagebox.showerror("로드 오류", str(e))

    def _save_excel_as(self):
        """다른 이름으로 저장 — 항상 다이얼로그 표시"""
        init_name = os.path.basename(self._excel_manager.filepath) \
            if self._excel_manager.filepath else "groups.xlsx"

        path = filedialog.asksaveasfilename(
            title="다른 이름으로 저장",
            defaultextension=".xlsx",
            filetypes=[("Excel 파일", "*.xlsx")],
            initialfile=init_name
        )
        if not path:
            return
        self._excel_manager.filepath = path
        self._do_save(path)

    def _save_excel_current(self):
        """현재 파일에 저장 — 파일 없으면 다이얼로그"""
        if not self._excel_manager.filepath:
            self._save_excel_as()
            return
        self._do_save(self._excel_manager.filepath)

    def _do_save(self, path: str):
        try:
            self._excel_manager.save_all(self._rows)
            self._update_filepath_label(path)
            self._log.append(fmt_log(
                f"📤 저장 완료: {os.path.basename(path)}"))
            self._log.append(fmt_log(f"   경로: {path}"))
            messagebox.showinfo(
                "저장 완료",
                f"✅ 저장되었습니다.\n\n경로: {path}"
            )
        except Exception as e:
            messagebox.showerror("저장 오류", str(e))

    def _refresh(self):
        if self._excel_manager.filepath:
            try:
                self._rows = self._excel_manager.load()
            except Exception:
                pass
        self._apply_filter()
        self._update_stats()

    def _apply_filter(self, *args):
        status_f = self._filter_var.get()
        search_q = self._search_var.get().lower()

        filtered = []
        for row in self._rows:
            if status_f != "전체" and row.status != status_f:
                continue
            if search_q and search_q not in row.link.lower():
                continue
            filtered.append(row)

        self._filtered_rows = filtered
        self._refresh_tree()

    def _refresh_tree(self):
        self._tree.delete(*self._tree.get_children())
        for i, row in enumerate(self._filtered_rows, start=1):
            tag = {
                "가입완료": "joined",
                "발송완료": "sent",
                "실패":    "failed",
                "건너뜀":  "skipped",
            }.get(row.status, "default")

            self._tree.insert("", "end", values=(
                i, row.link, row.status,
                row.processed_at or "-",
                row.note or "-"
            ), tags=(tag,))

    def _update_stats(self):
        stats = self._excel_manager.get_stats()
        for key, lbl in self._stat_labels.items():
            lbl.config(text=str(stats.get(key, 0)))

    def _on_double_click(self, event):
        item = self._tree.identify_row(event.y)
        col  = self._tree.identify_column(event.x)
        if not item or col != "#3":
            return
        idx = int(self._tree.item(item, "values")[0]) - 1
        if not (0 <= idx < len(self._filtered_rows)):
            return

        row = self._filtered_rows[idx]
        statuses = ["미가입", "가입완료", "발송완료", "실패", "건너뜀"]
        cur = row.status if row.status in statuses else "미가입"
        nxt = statuses[(statuses.index(cur) + 1) % len(statuses)]
        row.status = nxt

        if self._excel_manager.filepath:
            self._excel_manager.update_row(row.row_index, nxt)
        self._apply_filter()
        self._update_stats()

    def _bulk_change_status(self):
        selected = self._tree.selection()
        if not selected:
            messagebox.showwarning("경고", "선택된 항목이 없습니다.")
            return
        new_st = self._new_status_var.get()
        for item in selected:
            idx = int(self._tree.item(item, "values")[0]) - 1
            if 0 <= idx < len(self._filtered_rows):
                row = self._filtered_rows[idx]
                row.status = new_st
                if self._excel_manager.filepath:
                    self._excel_manager.update_row(row.row_index, new_st)
        self._apply_filter()
        self._update_stats()
        self._log.append(fmt_log(f"✏️ {len(selected)}개 → {new_st}"))

    def _delete_selected(self):
        selected = self._tree.selection()
        if not selected:
            messagebox.showwarning("경고", "선택된 항목이 없습니다.")
            return
        if not messagebox.askyesno("삭제 확인",
                                    f"{len(selected)}개 항목을 삭제할까요?"):
            return
        to_del = set()
        for item in selected:
            idx = int(self._tree.item(item, "values")[0]) - 1
            if 0 <= idx < len(self._filtered_rows):
                to_del.add(self._filtered_rows[idx].link)
        self._rows = [r for r in self._rows if r.link not in to_del]
        self._excel_manager._rows = self._rows
        self._apply_filter()
        self._update_stats()
        self._log.append(fmt_log(f"🗑️ {len(to_del)}개 삭제"))

    def _remove_duplicates(self):
        seen, unique, removed = set(), [], 0
        for row in self._rows:
            key = row.link.lower()
            if key not in seen:
                seen.add(key)
                unique.append(row)
            else:
                removed += 1
        self._rows = unique
        self._excel_manager._rows = unique
        self._apply_filter()
        self._update_stats()
        self._log.append(fmt_log(f"🔗 중복 {removed}개 제거"))

    def _copy_all_links(self):
        import pyperclip
        links = "\n".join(r.link for r in self._filtered_rows)
        pyperclip.copy(links)
        self._log.append(fmt_log(
            f"📋 {len(self._filtered_rows)}개 링크 복사"))

    def _add_links(self):
        raw = self._add_text.get("1.0", "end")
        stats = count_stats(raw)
        new_links = parse_links(raw)
        if not new_links:
            messagebox.showinfo("알림", "유효한 링크가 없습니다.")
            return

        existing = {r.link.lower() for r in self._rows}
        added = 0
        max_row = max((r.row_index for r in self._rows), default=1)

        for lk in new_links:
            if lk.lower() not in existing:
                max_row += 1
                self._rows.append(GroupRow(row_index=max_row, link=lk))
                existing.add(lk.lower())
                added += 1

        self._excel_manager._rows = self._rows
        self._add_text.delete("1.0", "end")
        self._apply_filter()
        self._update_stats()
        self._log.append(fmt_log(
            f"➕ {added}개 추가 (중복 {stats['parsed'] - added}개 스킵)"))


# ════════════════════════════════════════════════════════════
# [E] Entry Layer — TelegramAllInOne v2.0 + main()
# ════════════════════════════════════════════════════════════

# ── E1: main_window.py (v2.0) ───────────────────────
# [DUP_REMOVED] import tkinter as tk
# [DUP_CLEAN] from tkinter import ttk
# [DUP_REMOVED] import keyboard

# [MERGED] import config as cfg
# [DUP_CLEAN] from ui.coord_tab import CoordTab
# [DUP_CLEAN] from ui.join_tab import JoinTab
# [DUP_CLEAN] from ui.join_only_tab import JoinOnlyTab
# [DUP_CLEAN] from ui.message_tab import MessageTab
# [DUP_CLEAN] from ui.manager_tab import ManagerTab


class TelegramAllInOne:
    """텔레그램 올인원 자동화 메인 창"""

    def __init__(self, root: tk.Tk):
        self.root = root
        self._setup_window()
        self._build_ui()
        self._register_hotkeys()

    # ─── 창 설정 ─────────────────────────────────────────
    def _setup_window(self):
        self.root.title("🚀 텔레그램 올인원 자동화 v2.0")
        self.root.geometry("1200x860")
        self.root.minsize(900, 640)
        self.root.configure(bg=THEME["bg"])

        # 닫기 시 정리
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # ttk 스타일
        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure("TNotebook",
                        background=THEME["bg"],
                        borderwidth=0)
        style.configure("TNotebook.Tab",
                        background=THEME["card"],
                        foreground=THEME["subtext"],
                        font=FONT["header"],
                        padding=[16, 8])
        style.map("TNotebook.Tab",
                  background=[("selected", THEME["border"])],
                  foreground=[("selected", THEME["text"])])
        style.configure("TProgressbar",
                        background=THEME["accent"],
                        troughcolor=THEME["input_bg"])
        style.configure("Vertical.TScrollbar",
                        background=THEME["border"],
                        troughcolor=THEME["input_bg"])
        style.configure("TCombobox",
                        fieldbackground=THEME["input_bg"],
                        background=THEME["border"],
                        foreground=THEME["text"])

    # ─── UI 구성 ─────────────────────────────────────────
    def _build_ui(self):
        # 헤더
        self._build_header()

        # 노트북 (탭)
        self._notebook = ttk.Notebook(self.root)
        self._notebook.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        self._notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

        # 탭 생성
        self.coord_tab = CoordTab(self._notebook)
        self.join_only_tab = JoinOnlyTab(
            self._notebook,
            get_coord_config=self.coord_tab.get_config
        )
        self.join_tab = JoinTab(
            self._notebook,
            get_coord_config=self.coord_tab.get_config
        )
        self.message_tab = MessageTab(
            self._notebook,
            get_coord_config=self.coord_tab.get_config
        )
        self.manager_tab = ManagerTab(self._notebook)

        # 탭 순서: 좌표설정(0) / 가입전용(1) / 자동합류+메시지(2) / 메시지발송(3) / 그룹관리(4)
        self._notebook.add(self.coord_tab,     text="📍 좌표 설정")
        self._notebook.add(self.join_only_tab, text="🎯 가입 전용")
        self._notebook.add(self.join_tab,      text="🚪 합류+메시지")
        self._notebook.add(self.message_tab,   text="💬 메시지 발송")
        self._notebook.add(self.manager_tab,   text="📊 그룹 관리")

        # 상태바
        self._build_statusbar()

    def _build_header(self):
        header = tk.Frame(self.root, bg=THEME["card"], height=52)
        header.pack(fill="x", padx=0)
        header.pack_propagate(False)

        # 타이틀
        tk.Label(
            header,
            text="🚀 텔레그램 올인원 자동화 v2.0",
            bg=THEME["card"],
            fg=THEME["accent"],
            font=("맑은 고딕", 15, "bold")
        ).pack(side="left", padx=20, pady=10)

        # 핫키 안내
        hint_text = "F8: 시작   F9: 중단   Ctrl+1~5: 탭 전환"
        tk.Label(
            header,
            text=hint_text,
            bg=THEME["card"],
            fg=THEME["subtext"],
            font=FONT["small"]
        ).pack(side="right", padx=20)

    def _build_statusbar(self):
        bar = tk.Frame(self.root, bg=THEME["card"], height=26)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)

        self._status_var = tk.StringVar(value="준비")
        tk.Label(
            bar,
            textvariable=self._status_var,
            bg=THEME["card"],
            fg=THEME["subtext"],
            font=FONT["small"],
            anchor="w"
        ).pack(side="left", padx=12)

        # pyautogui FAILSAFE 안내
        tk.Label(
            bar,
            text="⚠️ 마우스를 화면 좌상단으로 이동하면 긴급 중단",
            bg=THEME["card"],
            fg=THEME["warning"],
            font=FONT["small"]
        ).pack(side="right", padx=12)

    # ─── 핫키 ────────────────────────────────────────────
    def _register_hotkeys(self):
        try:
            keyboard.add_hotkey("f8", self._hotkey_start)
            keyboard.add_hotkey("f9", self._hotkey_stop)
            keyboard.add_hotkey("ctrl+1", lambda: self._switch_tab(0))
            keyboard.add_hotkey("ctrl+2", lambda: self._switch_tab(1))
            keyboard.add_hotkey("ctrl+3", lambda: self._switch_tab(2))
            keyboard.add_hotkey("ctrl+4", lambda: self._switch_tab(3))
            keyboard.add_hotkey("ctrl+5", lambda: self._switch_tab(4))
        except Exception:
            pass

    def _unregister_hotkeys(self):
        try:
            keyboard.unhook_all_hotkeys()
        except Exception:
            pass

    def _hotkey_start(self):
        tab_idx = self._notebook.index("current")
        if tab_idx == 1:
            self.root.after(0, self.join_only_tab.start)
        elif tab_idx == 2:
            self.root.after(0, self.join_tab.start)
        elif tab_idx == 3:
            self.root.after(0, self.message_tab.start)

    def _hotkey_stop(self):
        tab_idx = self._notebook.index("current")
        if tab_idx == 1:
            self.root.after(0, self.join_only_tab.stop)
        elif tab_idx == 2:
            self.root.after(0, self.join_tab.stop)
        elif tab_idx == 3:
            self.root.after(0, self.message_tab.stop)

    def _switch_tab(self, idx: int):
        try:
            self._notebook.select(idx)
        except Exception:
            pass

    # ─── 탭 변경 이벤트 ──────────────────────────────────
    def _on_tab_change(self, event=None):
        tab_names = ["좌표 설정", "가입 전용", "합류+메시지", "메시지 발송", "그룹 관리"]
        try:
            idx = self._notebook.index("current")
            self._status_var.set(f"현재 탭: {tab_names[idx]}")
        except Exception:
            pass

    # ─── 종료 ────────────────────────────────────────────
    def _on_close(self):
        # 실행 중인 작업 중단
        from utils.stop_manager import stop_controller
        stop_controller.stop_all()

        # 핫키 해제
        self._unregister_hotkeys()

        # 좌표 설정 자동 저장
        try:
            self.coord_tab._save_config()
        except Exception:
            pass

        self.root.destroy()


# ── E2: main() 진입점 ────────────────────────────────
# [DUP_REMOVED] import sys
# [DUP_REMOVED] import os

# 프로젝트 루트를 Python 경로에 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# [DUP_REMOVED] import tkinter as tk
# [DUP_REMOVED] import pyautogui

# [DUP_CLEAN] from ui.main_window import TelegramAllInOne
# [MERGED] import config as cfg


def main():
    # ─── pyautogui 설정 ──────────────────────────────
    pyautogui.FAILSAFE = True   # 좌상단 이동 시 긴급 중단
    pyautogui.PAUSE = 0         # 자동 딜레이 없음 (수동 관리)

    # ─── 필수 폴더 생성 ──────────────────────────────
    os.makedirs(get_presets_dir("coord"), exist_ok=True)
    os.makedirs(get_outputs_dir(), exist_ok=True)

    # ─── Tkinter 메인 창 ─────────────────────────────
    root = tk.Tk()
    TelegramAllInOne(root)  # instance kept alive by root.mainloop()

    try:
        root.mainloop()
    except KeyboardInterrupt:
        pass
    except Exception as e:
        print(f"[오류] {e}")
    finally:
        try:
            import keyboard
            keyboard.unhook_all_hotkeys()
        except Exception:
            pass


if __name__ == "__main__":
    main()