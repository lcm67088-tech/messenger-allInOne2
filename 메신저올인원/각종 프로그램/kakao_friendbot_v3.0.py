#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
카카오톡 친구추가 봇 PRO v3.0
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
단일 파이썬 파일 통합본  (kakao_friendbot_v3.0.py)

[v3.0 주요 변경사항]
  • 신규 친구 추가 재시도 로직 (최대 N회)
  • 이어하기(Resume) 모드 – 중단 위치부터 재개
  • 설정 자동저장 (프로그램 종료 시)
  • ETA(남은 시간) 실시간 표시
  • 프리셋 삭제 기능
  • 실행 로그 파일 저장 (logs/)
  • Excel 결과에 요약 시트 자동 추가
  • v3.0 옵션 패널 (재시도/이어하기/로그파일)
  • 다크 실시간 로그 (키워드별 색상)
  • 결과 트리 행 색상 태그
  • 현재 처리 ID 전용 표시 라벨
  • 패키지 상태 확인 메뉴
  • OCR 드래그 영역: Esc 취소 지원
  • OCR 좌표 자동 정합성 보정 (좌상단/우하단 순서)
  • 이미지 전처리 스케일업 (소자 인식률 향상)
  • 마우스 휠 스크롤 지원 (설정 패널)

[통합 시 충돌 해결 목록]
  - E1/E2의 parts.* import → 통합 파일에서 직접 함수/클래스 참조로 전환
  - F3의 parts.A2_ocr_engine import → 동일 파일 내 함수 직접 호출
  - C1의 parts.A1_foundation import → 통합 파일 내 함수 직접 호출
  - 모든 믹스인 import 구조 제거, 단일 클래스 KakaoFriendBotPro로 통합
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""

# ============================================================
# 0. 표준 라이브러리 임포트
# ============================================================
import os
import sys
import re
import json
import time
import threading
import difflib
import subprocess
from datetime import datetime, timedelta

# ============================================================
# 1. GUI 임포트
# ============================================================
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog

# ============================================================
# 2. 서드파티 라이브러리 임포트
# ============================================================
import pyautogui
import pyperclip
import pytesseract
from PIL import Image, ImageGrab, ImageOps, ImageFilter, ImageEnhance

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
import keyboard

# ============================================================
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# [A-1] 기반 인프라: 경로/환경 유틸 + Tesseract + 패키지체커
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ============================================================

def _norm(p: str) -> str:
    return os.path.normpath((p or "").strip().strip('"').strip("'"))

def _app_dir() -> str:
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

def _resource_path(rel_path: str) -> str:
    return os.path.join(_app_dir(), rel_path)

_BASE_DIR = _norm(_app_dir())
_TESS_DIR = _norm(os.path.join(_BASE_DIR, "tesseract"))
_TESS_EXE = _norm(os.path.join(_TESS_DIR, "tesseract.exe"))
_TESSDATA = _norm(os.path.join(_TESS_DIR, "tessdata"))

def _maybe_set_portable_tesseract() -> bool:
    try:
        if os.path.isfile(_TESS_EXE):
            pytesseract.pytesseract.tesseract_cmd = _TESS_EXE
        if os.path.isdir(_TESSDATA):
            os.environ["TESSDATA_PREFIX"] = _TESSDATA
        return os.path.isfile(_TESS_EXE) and os.path.isdir(_TESSDATA)
    except Exception:
        return False

def get_dependency_status_text() -> str:
    pkgs = ["pyautogui", "pyperclip", "pytesseract", "PIL", "pandas",
            "openpyxl", "keyboard"]
    lines = ["=== 의존성 패키지 현황 ==="]
    for pkg in pkgs:
        try:
            __import__(pkg)
            lines.append(f"  ✅ {pkg}")
        except ImportError:
            lines.append(f"  ❌ {pkg} (미설치)")
    try:
        ver = pytesseract.get_tesseract_version()
        lines.append(f"\n  Tesseract: {ver}")
    except Exception:
        lines.append("\n  Tesseract: ❌ 미설치/미설정")
    lines.append(f"  TESSDATA_PREFIX: {os.environ.get('TESSDATA_PREFIX','(미설정)')}")
    return "\n".join(lines)

# 모듈 로드 시 포터블 Tesseract 우선 적용
_maybe_set_portable_tesseract()
if os.path.isdir(_TESSDATA):
    print(f"✅ TESSDATA_PREFIX: {os.environ.get('TESSDATA_PREFIX','(미설정)')}")
else:
    print(f"⚠️  tessdata 폴더 없음: {_TESSDATA}")
if os.path.isfile(_TESS_EXE):
    print(f"✅ Tesseract EXE: {_TESS_EXE}")
else:
    print(f"⚠️  Tesseract EXE 없음: {_TESS_EXE}")


# ============================================================
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# [A-2] OCR 엔진
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ============================================================

def preprocess_for_ocr(img: Image.Image) -> Image.Image:
    gray = ImageOps.grayscale(img)
    gray = ImageEnhance.Contrast(gray).enhance(2.0)
    gray = gray.filter(ImageFilter.GaussianBlur(radius=0.5))
    bw   = gray.point(lambda p: 255 if p > 155 else 0, mode='1')
    bw_l = bw.convert('L')
    w, h = bw_l.size
    if w < 200 or h < 50:
        bw_l = bw_l.resize((w * 2, h * 2), Image.LANCZOS)
    return bw_l

def clean_text(s: str) -> str:
    return (s or "").replace(' ', '').replace('\n', '').replace('\t', '').lower()

_OCR_FIX_MAP = {
    '친구주가':'친구추가','친구추거':'친구추가','친구추갸':'친구추가',
    '친구추가하긱':'친구추가하기','친구추기':'친구추가','친구추값':'친구추가',
    '프르필':'프로필','프리필':'프로필','프료필':'프로필',
    '체팅':'채팅','채텽':'채팅','챈널':'채널','쳔널':'채널',
    '찿을':'찾을','없섯':'없었','없슴':'없음',
    '111':'1:1','차딘':'차단','자단':'차단',
    '채 널':'채널','채 팅':'채팅',
    '진구추가':'친구추가','찬구추가':'친구추가','천구추가':'친구추가',
    '친구츄가':'친구추가','친구쵸가':'친구추가',
    '채딩':'채팅','차팅':'채팅','차딩':'채팅','채링':'채팅',
    '체널':'채널','채녈':'채널','표로필':'프로필','프로빌':'프로필',
    '업섯':'없었','업슴':'없음','1ㅣ1':'1:1','ㅣㅣㅣ':'1:1',
    '친구 추가':'친구추가',
    '친구추가하키':'친구추가하기','친구추가히기':'친구추가하기',
    '친구추다':'친구추가','친구추고':'친구추가',
    '채뎡':'채팅','채텅':'채팅',
    '없어':'없음','없이':'없음',
}

_VARIANT_GROUPS = [
    (r'(진구|친그|찬구|찬그|천구|천그|칠구|친규)', '친구'),
    (r'(주가|주기|추기|츠가|츠기|츄가|쵸가|추거|주거|추다|추고)', '추가'),
    (r'(채딩|차팅|차딩|채링|차링|체팅|챠팅|채뎡|채텅)', '채팅'),
    (r'(챈널|쳔널|체널|채녈|챤널)', '채널'),
    (r'(프르필|프리필|프료필|표로필|프로빌|프로핀|프로팔)', '프로필'),
    (r'(없슴|업슴|없섯|업섯|읍슴|읍섯|없어|없이)', '없'),
    (r'(찿을|차을|착을)', '찾을'),
]

def apply_variant_groups(text: str) -> str:
    s = text or ""
    for pat, rep in _VARIANT_GROUPS:
        s = re.sub(pat, rep, s)
    return s

def _bigrams(s: str) -> set:
    return set(s[i:i+2] for i in range(len(s)-1)) if len(s) >= 2 else {s}

def _sim_ratio_jaccard(a: str, b: str) -> float:
    if not a and not b: return 1.0
    A, B = _bigrams(a), _bigrams(b)
    if not A or not B: return 0.0
    return len(A & B) / max(1, len(A | B))

def _sim_ratio_seq(a: str, b: str) -> float:
    return difflib.SequenceMatcher(None, a, b).ratio()

def any_token_like(text: str, tokens, thresh_j=0.40, thresh_seq=0.60) -> bool:
    if not text: return False
    for t in tokens:
        t = clean_text(t)
        if not t: continue
        if t in text or text in t: return True
        if _sim_ratio_jaccard(text, t) >= thresh_j: return True
        if _sim_ratio_seq(text, t)     >= thresh_seq: return True
    return False

TOK_NEW = [
    '친구추가','친구추가하기','추가','새친구',
    '친구주가','진구추가','찬구추가','천구추가',
    '친구츄가','친구쵸가','친구추기','친구주기','친구추다','친구추고',
]
TOK_EXISTING = [
    '채팅','1:1','프로필','멀티프로필','멀티','대화상대',
    '채딩','차팅','체팅','챠팅','111','1ㅣ1',
    '프르필','프리필','프료필','표로필','채뎡','채텅',
]
TOK_CHANNEL  = ['채널','채널추가','챈널','쳔널','체널','채녈']
TOK_NOTFOUND = [
    '찾을','없음','없었','검색결과없음','검색결과가없습니다',
    '찿을','없슴','없섯','업슴','업섯','없어','없이',
]

def decide_state_from_ocr(clean_text_result: str) -> str:
    txt = clean_text_result
    if any_token_like(txt, TOK_NEW):      return '신규'
    if any_token_like(txt, TOK_EXISTING): return '기존'
    if any_token_like(txt, TOK_CHANNEL):  return '채널'
    if any_token_like(txt, TOK_NOTFOUND) or len(txt) < 3: return '없음'
    return '오류'

def ocr_text_strong(img_raw: Image.Image) -> tuple:
    proc = preprocess_for_ocr(img_raw)
    tess_cmd  = getattr(pytesseract.pytesseract, 'tesseract_cmd', '') or 'tesseract'
    tess_base = os.path.dirname(tess_cmd)
    tessdata_dir = os.path.normpath(os.path.join(tess_base, 'tessdata'))
    env_td = os.environ.get("TESSDATA_PREFIX", "")
    if env_td and os.path.isdir(env_td):
        tessdata_dir = env_td

    def _ocr(img):
        try:
            cfg = f'--tessdata-dir {tessdata_dir}' if tessdata_dir and os.path.isdir(tessdata_dir) else ''
            return pytesseract.image_to_string(img, lang='kor+eng', config=cfg) or ""
        except Exception as e:
            return f"(OCR오류:{e})"

    try:    text_raw  = _ocr(img_raw)
    except Exception as e: text_raw  = f"(원본OCR오류:{e})"
    try:    text_proc = _ocr(proc)
    except Exception as e: text_proc = f"(전처리OCR오류:{e})"

    use_text = text_proc if clean_text(text_proc) else text_raw
    fixed = use_text
    for k, v in _OCR_FIX_MAP.items():
        fixed = fixed.replace(k, v)
    fixed = apply_variant_groups(fixed)
    return clean_text(fixed), text_raw, text_proc, proc


# ============================================================
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# [B-1] 설정 구조
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ============================================================

DEFAULT_CONFIG = {
    'version': '3.0',
    'coords': {
        'id_add_button':  None,
        'chat_button':    None,
        'profile_area':   None,
        'confirm_button': None,
        'close_button':   None,
    },
    'ocr_area': None,
    'timing': {
        'after_ctrlA': 2.0,
        'after_click': 1.5,
        'after_input': 2.5,
        'after_ocr':   3.0,
        'after_tab':   0.5,
    },
    'csv_path':           '',
    'output_path':        'run_output.xlsx',
    'counter_start':      1,
    'retry_count':        2,
    'auto_save_config':   True,
    'log_to_file':        False,
    'log_dir':            'logs',
    'last_processed_idx': 0,
    'resume_mode':        False,
}

def deep_merge_into(dst: dict, src: dict) -> dict:
    for k, v in src.items():
        if isinstance(v, dict) and isinstance(dst.get(k), dict):
            deep_merge_into(dst[k], v)
        else:
            dst[k] = v
    return dst

def make_config_copy() -> dict:
    return json.loads(json.dumps(DEFAULT_CONFIG))


# ============================================================
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# [B-2] CSV 처리
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ============================================================

def validate_csv_path(path: str) -> tuple:
    if not path:         return False, "CSV 경로가 비어있습니다."
    if not os.path.isfile(path): return False, f"CSV 파일이 존재하지 않습니다:\n{path}"
    if not path.lower().endswith(".csv"): return False, "CSV 파일(.csv)만 지원합니다."
    return True, "OK"

def read_csv_safely(path: str) -> 'pd.DataFrame':
    encodings = ["utf-8","cp949","utf-8-sig","euc-kr","latin-1"]
    delims    = [",",";","\t","|"]
    last_err  = None
    for enc in encodings:
        for sep in delims:
            try:
                df = pd.read_csv(path, encoding=enc, sep=sep, dtype=str)
                if df is not None and df.shape[1] >= 1 and df.shape[0] >= 1:
                    first_col = df.columns[0].strip().lower()
                    if first_col not in ('id','kakao_id','아이디','카카오아이디'):
                        df = pd.read_csv(path, encoding=enc, sep=sep, header=None, dtype=str)
                        df.columns = ['id'] + [f'col{i}' for i in range(1, len(df.columns))]
                    else:
                        df = df.rename(columns={df.columns[0]: 'id'})
                    return df
            except Exception as e:
                last_err = e
    raise RuntimeError(f"CSV 로딩 실패. 마지막 오류: {last_err}\n파일: {path}")


# ============================================================
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# [Main Class] KakaoFriendBotPro v3.0
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ============================================================

class KakaoFriendBotPro:
    """
    카카오톡 친구추가 봇 PRO v3.0
    모든 파트(A~F)를 단일 클래스로 통합.
    """

    def __init__(self, root: tk.Tk):
        # ── 상태 변수
        self.config           = make_config_copy()
        self.is_running       = False
        self.is_paused        = False
        self.current_id_index = 0
        self.stats            = {'신규': 0, '기존': 0, '없음': 0, '오류': 0}
        self._start_time      = 0.0

        # ── 경로 변수 (C4에서 참조)
        self.csv_path_var  = tk.StringVar(value="")
        self.output_dir_var = tk.StringVar(value="")
        self.xlsx_save_var  = tk.StringVar(value="")

        # ── 창 초기화
        self._init_window(root)

        # ── UI 빌드
        self._build_menu()
        self._build_layout()

        # ── 설정 자동 불러오기 (silent)
        self.load_config(silent=True)

        # ── 핫키 등록
        self._setup_hotkeys()

        # ── 경로 정리
        self.sanitize_paths(log=True)

    # ─────────────────────────────────────────────────────
    # [C-1] 메인창 초기화
    # ─────────────────────────────────────────────────────
    def _init_window(self, root: tk.Tk):
        self.root = root
        self.root.title("카카오톡 친구추가 봇 PRO v3.0")
        self.root.geometry("1180x800")
        self.root.minsize(900, 600)
        self.root.configure(bg='#f0f0f0')
        try:
            self.root.iconbitmap("kakao_icon.ico")
        except Exception:
            pass
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        self.auto_save_on_exit()
        try: keyboard.unhook_all()
        except Exception: pass
        self.root.destroy()

    # ─────────────────────────────────────────────────────
    # [C-1-2] 메뉴 바
    # ─────────────────────────────────────────────────────
    def _build_menu(self):
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="파일", menu=file_menu)
        file_menu.add_command(label="설정 저장",     command=self.save_config)
        file_menu.add_command(label="설정 불러오기",  command=self.load_config)
        file_menu.add_separator()
        file_menu.add_command(label="프리셋 저장",    command=self.save_preset)
        file_menu.add_command(label="프리셋 불러오기", command=self.load_preset)
        file_menu.add_separator()
        file_menu.add_command(label="종료",           command=self._on_close)

        run_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="실행", menu=run_menu)
        run_menu.add_command(label="▶ 시작  (F1)",    command=self.start_bot)
        run_menu.add_command(label="⏸ 일시정지 (F1)", command=self.pause_bot)
        run_menu.add_command(label="⏹ 중지  (F2)",    command=self.stop_bot)

        tool_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="도구", menu=tool_menu)
        tool_menu.add_command(label="OCR 테스트 (F10)",  command=self.ocr_test_capture_and_save)
        tool_menu.add_command(label="패키지 상태 확인",   command=self._show_dependency_status)
        tool_menu.add_separator()
        tool_menu.add_command(label="결과 로그 내보내기", command=self.export_log)
        tool_menu.add_command(label="결과 로그 새로고침", command=self.refresh_log)
        tool_menu.add_command(label="결과 로그 초기화",   command=self.clear_log)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="도움말", menu=help_menu)
        help_menu.add_command(label="핫키 안내",  command=self.show_hotkeys)
        help_menu.add_command(label="사용 설명서", command=self.show_manual)
        help_menu.add_separator()
        help_menu.add_command(label="버전 정보",   command=self._show_version_info)

    def _show_dependency_status(self):
        messagebox.showinfo("패키지 상태", get_dependency_status_text())

    def _show_version_info(self):
        messagebox.showinfo("버전 정보",
            "카카오톡 친구추가 봇 PRO\n버전: v3.0\n\n"
            "v3.0 주요 변경사항:\n"
            "  • 신규 친구 추가 재시도 로직\n  • 이어하기(Resume) 기능\n"
            "  • 설정 자동저장\n  • ETA 표시\n  • 요약 Excel 시트\n"
            "  • 다크 실시간 로그\n  • 프리셋 삭제 기능\n"
            "  • 패키지 상태 확인")

    # ─────────────────────────────────────────────────────
    # [C-1-3] 레이아웃 골격
    # ─────────────────────────────────────────────────────
    def _build_layout(self):
        main = ttk.Frame(self.root)
        main.pack(fill='both', expand=True, padx=6, pady=6)
        left  = ttk.LabelFrame(main, text="⚙️  설정",  padding=6)
        mid   = ttk.LabelFrame(main, text="▶️  실행",  padding=6)
        right = ttk.LabelFrame(main, text="📋  로그",  padding=6)
        left.grid( row=0, column=0, sticky='nsew', padx=(0, 4))
        mid.grid(  row=0, column=1, sticky='nsew', padx=4)
        right.grid(row=0, column=2, sticky='nsew', padx=(4, 0))
        main.columnconfigure(0, weight=25)
        main.columnconfigure(1, weight=12)
        main.columnconfigure(2, weight=20)
        main.rowconfigure(0, weight=1)
        self._create_left_panel(left)
        self._create_execute_panel(mid)
        self._create_right_panel(right)

    # ─────────────────────────────────────────────────────
    # [C-2~4] 왼쪽 패널
    # ─────────────────────────────────────────────────────
    def _create_left_panel(self, parent):
        canvas    = tk.Canvas(parent, bg='white', highlightthickness=0)
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        scrollable = ttk.Frame(canvas)
        scrollable.bind("<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        def _mw(event): canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _mw)
        self._create_section_config_mgmt(scrollable)
        self._create_section_coords(scrollable)
        self._create_section_ocr_area(scrollable)
        self._create_section_timing(scrollable)
        self._create_section_file(scrollable)
        self._create_section_v3_options(scrollable)
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def _create_section_config_mgmt(self, parent):
        f = ttk.LabelFrame(parent, text="💾  설정 관리", padding=10)
        f.grid(row=0, column=0, padx=10, pady=6, sticky='ew')
        f.columnconfigure((0,1,2,3), weight=1)
        for col, (lbl, cmd) in enumerate([
            ("설정 저장", self.save_config), ("설정 불러오기", self.load_config),
            ("프리셋 저장", self.save_preset), ("프리셋 불러오기", self.load_preset)
        ]):
            ttk.Button(f, text=lbl, command=cmd).grid(row=0, column=col, padx=3, pady=3, sticky='ew')

    def _create_section_coords(self, parent):
        f = ttk.LabelFrame(parent, text="🎯  좌표 설정 (핫키 F3~F7)", padding=10)
        f.grid(row=1, column=0, padx=10, pady=6, sticky='ew')
        f.columnconfigure(1, weight=1)
        ttk.Label(f, text="버튼 위에 마우스를 올린 뒤 해당 핫키를 누르세요.",
                  foreground='#1565c0').grid(row=0, column=0, columnspan=3, pady=(0,8), sticky='w')
        coord_defs = [
            ('id_add_button', 'ID로 추가',  'F3'),
            ('chat_button',   '1:1 채팅',   'F4'),
            ('profile_area',  '프로필 영역', 'F5'),
            ('confirm_button','확인',        'F6'),
            ('close_button',  'X 닫기',     'F7'),
        ]
        self.coord_labels = {}
        for i, (key, name, hk) in enumerate(coord_defs, 1):
            ttk.Label(f, text=f"[{hk}] {name}:").grid(row=i, column=0, sticky='w', padx=(5,2), pady=3)
            lbl = tk.Label(f, text="미설정", bg='#ffebee', fg='#c62828',
                           relief='groove', width=22, font=('맑은 고딕', 9))
            lbl.grid(row=i, column=1, padx=5, pady=3, sticky='ew')
            self.coord_labels[key] = lbl
            ttk.Button(f, text="📍 3초 후 캡처",
                       command=lambda k=key: self.capture_coord(k), width=14)\
                .grid(row=i, column=2, padx=5, pady=3)

    def _create_section_ocr_area(self, parent):
        f = ttk.LabelFrame(parent, text="📷  OCR 영역 (F8/F9 또는 드래그)", padding=10)
        f.grid(row=2, column=0, padx=10, pady=6, sticky='ew')
        f.columnconfigure(1, weight=1)
        ttk.Label(f, text="좌상단(F8) 설정 후 → 우하단(F9) 설정",
                  foreground='#1565c0').grid(row=0, column=0, columnspan=3, pady=(0,6), sticky='w')
        ttk.Label(f, text="[F8] 좌상단:").grid(row=1, column=0, sticky='w', padx=5)
        self.ocr_topleft_label = tk.Label(f, text="미설정", bg='#ffebee', fg='#c62828',
                                          relief='groove', width=18, font=('맑은 고딕', 9))
        self.ocr_topleft_label.grid(row=1, column=1, padx=5, pady=3, sticky='ew')
        ttk.Label(f, text="[F9] 우하단:").grid(row=2, column=0, sticky='w', padx=5)
        self.ocr_bottomright_label = tk.Label(f, text="미설정", bg='#ffebee', fg='#c62828',
                                               relief='groove', width=18, font=('맑은 고딕', 9))
        self.ocr_bottomright_label.grid(row=2, column=1, padx=5, pady=3, sticky='ew')
        ttk.Button(f, text="🖱️ 드래그", command=self.capture_ocr_area, width=10)\
            .grid(row=1, column=2, rowspan=2, padx=6, pady=3)
        ttk.Button(f, text="📸 OCR 테스트 캡처 → 저장 → 텍스트  (F10)",
                   command=self.ocr_test_capture_and_save)\
            .grid(row=3, column=0, columnspan=3, pady=(8,2), padx=5, sticky='ew')

    def _create_section_timing(self, parent):
        f = ttk.LabelFrame(parent, text="⏱️  타이밍 (초)", padding=10)
        f.grid(row=3, column=0, padx=10, pady=6, sticky='ew')
        timing_defs = [
            ('after_ctrlA','Ctrl+A 후', 0.01,10.0),
            ('after_click','클릭 후',   0.01,10.0),
            ('after_input','입력 후',   0.01,10.0),
            ('after_ocr',  'OCR 전',    0.01,10.0),
            ('after_tab',  'Tab 후',    0.01, 5.0),
        ]
        self.timing_vars = {}
        for i, (key, name, mn, mx) in enumerate(timing_defs):
            r, c = i//2, (i%2)*3
            ttk.Label(f, text=f"{name}:").grid(row=r, column=c, sticky='w', padx=(5,2), pady=3)
            var = tk.DoubleVar(value=self.config['timing'].get(key, 1.0))
            self.timing_vars[key] = var
            ttk.Spinbox(f, from_=mn, to=mx, textvariable=var,
                        width=8, increment=0.1, format="%.2f")\
                .grid(row=r, column=c+1, padx=(2,10), pady=3)

    def _create_section_file(self, parent):
        f = ttk.LabelFrame(parent, text="📁  파일 / 번호", padding=10)
        f.grid(row=4, column=0, padx=10, pady=6, sticky='ew')
        f.columnconfigure(1, weight=1)
        ttk.Label(f, text="입력 CSV:").grid(row=0, column=0, sticky='w', padx=5, pady=3)
        self.csv_entry = ttk.Entry(f, textvariable=self.csv_path_var, width=34)
        self.csv_entry.grid(row=0, column=1, padx=5, pady=3, sticky='ew')
        ttk.Button(f, text="찾기", command=self.browse_csv, width=7).grid(row=0, column=2, padx=3)
        ttk.Label(f, text="출력 폴더:").grid(row=1, column=0, sticky='w', padx=5, pady=3)
        self.output_entry = ttk.Entry(f, textvariable=self.output_dir_var, width=34)
        self.output_entry.grid(row=1, column=1, padx=5, pady=3, sticky='ew')
        ttk.Button(f, text="찾기", command=self.browse_output, width=7).grid(row=1, column=2, padx=3)
        ttk.Label(f, text="가망 번호:").grid(row=2, column=0, sticky='w', padx=5, pady=3)
        self.counter_var = tk.IntVar(value=self.config.get('counter_start', 1))
        ttk.Spinbox(f, from_=1, to=99999, textvariable=self.counter_var, width=12)\
            .grid(row=2, column=1, sticky='w', padx=5, pady=3)

    def browse_csv(self):
        path = filedialog.askopenfilename(
            title="CSV 파일 선택",
            filetypes=[("CSV files","*.csv"),("All files","*.*")])
        if path:
            self.csv_path_var.set(path)
            self.config["csv_path"] = path
            self.log_message(f"[파일] CSV 선택: {path}")

    def browse_output(self):
        path = filedialog.askdirectory(title="출력 폴더 선택")
        if path:
            self.output_dir_var.set(path)
            self.config["output_path"] = path
            self.log_message(f"[폴더] 출력 경로: {path}")

    def _create_section_v3_options(self, parent):
        f = ttk.LabelFrame(parent, text="🆕  v3.0 옵션", padding=10)
        f.grid(row=5, column=0, padx=10, pady=6, sticky='ew')
        f.columnconfigure(1, weight=1)
        ttk.Label(f, text="친구추가 재시도:").grid(row=0, column=0, sticky='w', padx=5, pady=4)
        self.retry_count_var = tk.IntVar(value=self.config.get('retry_count', 2))
        ttk.Spinbox(f, from_=0, to=5, textvariable=self.retry_count_var, width=6)\
            .grid(row=0, column=1, sticky='w', padx=5, pady=4)
        ttk.Label(f, text="회", foreground='gray').grid(row=0, column=2, sticky='w')
        self.resume_mode_var = tk.BooleanVar(value=self.config.get('resume_mode', False))
        ttk.Checkbutton(f, text="이어하기 모드 (마지막 위치부터 재개)",
                        variable=self.resume_mode_var)\
            .grid(row=1, column=0, columnspan=3, sticky='w', padx=5, pady=4)
        self.log_to_file_var = tk.BooleanVar(value=self.config.get('log_to_file', False))
        ttk.Checkbutton(f, text="실행 로그 파일 저장 (logs/ 폴더)",
                        variable=self.log_to_file_var)\
            .grid(row=2, column=0, columnspan=3, sticky='w', padx=5, pady=4)
        self.auto_save_var = tk.BooleanVar(value=self.config.get('auto_save_config', True))
        ttk.Checkbutton(f, text="종료 시 설정 자동저장",
                        variable=self.auto_save_var)\
            .grid(row=3, column=0, columnspan=3, sticky='w', padx=5, pady=4)

    # ─────────────────────────────────────────────────────
    # [C-5] 실행 패널
    # ─────────────────────────────────────────────────────
    def _create_execute_panel(self, parent):
        # 제어 버튼
        ctrl = ttk.LabelFrame(parent, text="🎮  제어", padding=10)
        ctrl.pack(fill='x', pady=(0,6))
        ttk.Label(ctrl, text="F1: 시작/재개  │  F2: 중지",
                  foreground='#555555', font=('맑은 고딕',8)).pack(pady=(0,6))
        self.start_btn = ttk.Button(ctrl, text="▶️  시작",      command=self.start_bot,  width=20)
        self.start_btn.pack(pady=2, fill='x')
        self.pause_btn = ttk.Button(ctrl, text="⏸️  일시정지", command=self.pause_bot,  width=20, state='disabled')
        self.pause_btn.pack(pady=2, fill='x')
        self.stop_btn  = ttk.Button(ctrl, text="⏹️  중지",     command=self.stop_bot,   width=20, state='disabled')
        self.stop_btn.pack(pady=2, fill='x')
        # 진행
        prog = ttk.LabelFrame(parent, text="📊  진행", padding=10)
        prog.pack(fill='x', pady=(0,6))
        self.progress_label = ttk.Label(prog, text="대기 중...", font=('맑은 고딕',9))
        self.progress_label.pack(pady=(0,4))
        self.progress_bar = ttk.Progressbar(prog, mode='determinate', length=200)
        self.progress_bar.pack(fill='x', pady=(0,2))
        self.eta_label = ttk.Label(prog, text="남은 시간: -", foreground='#777777', font=('맑은 고딕',8))
        self.eta_label.pack()
        # 통계
        stats = ttk.LabelFrame(parent, text="📈  통계", padding=10)
        stats.pack(fill='both', expand=True, pady=(0,6))
        self.stat_labels = {}
        for key, color in [('신규','#4caf50'),('기존','#2196f3'),('없음','#ff9800'),('오류','#f44336')]:
            fr = tk.Frame(stats, bg=color, relief='raised', bd=2)
            fr.pack(fill='both', expand=True, pady=2)
            tk.Label(fr, text=key, bg=color, fg='white', font=('맑은 고딕',9,'bold')).pack()
            lbl = tk.Label(fr, text="0", bg=color, fg='white', font=('맑은 고딕',16,'bold'))
            lbl.pack()
            self.stat_labels[key] = lbl
        # 현재 처리 ID
        cur = ttk.LabelFrame(parent, text="🔍  현재 처리 ID", padding=6)
        cur.pack(fill='x')
        self.current_id_label = tk.Label(cur, text="-", bg='#fafafa', fg='#333333',
                                         relief='sunken', font=('Consolas',10,'bold'), anchor='center')
        self.current_id_label.pack(fill='x', pady=2, ipady=4)

    # ─────────────────────────────────────────────────────
    # [C-6] 오른쪽 패널
    # ─────────────────────────────────────────────────────
    def _create_right_panel(self, parent):
        # 실시간 로그
        realtime = ttk.LabelFrame(parent, text="📝  실시간 로그", padding=5)
        realtime.pack(fill='both', expand=True, pady=(0,5))
        self.realtime_log = scrolledtext.ScrolledText(
            realtime, height=14, font=('Consolas',8),
            state='disabled', wrap='word',
            bg='#1e1e1e', fg='#d4d4d4', insertbackground='white')
        self.realtime_log.pack(fill='both', expand=True)
        self.realtime_log.tag_config('ok',    foreground='#6dc05e')
        self.realtime_log.tag_config('warn',  foreground='#f0c050')
        self.realtime_log.tag_config('error', foreground='#f05050')
        self.realtime_log.tag_config('info',  foreground='#9cdcfe')
        # 결과 트리
        result = ttk.LabelFrame(parent, text="📋  결과 로그", padding=5)
        result.pack(fill='both', expand=True)
        columns = ('ID','결과','이름','시각')
        self.log_tree = ttk.Treeview(result, columns=columns, show='headings', height=8)
        col_w = {'ID':130,'결과':60,'이름':80,'시각':130}
        for col in columns:
            self.log_tree.heading(col, text=col)
            self.log_tree.column(col, width=col_w.get(col,100), anchor='center')
        vsb = ttk.Scrollbar(result, orient="vertical",   command=self.log_tree.yview)
        hsb = ttk.Scrollbar(result, orient="horizontal", command=self.log_tree.xview)
        self.log_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.log_tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        result.rowconfigure(0, weight=1); result.columnconfigure(0, weight=1)
        self.log_tree.tag_configure('신규', background='#e8f5e9')
        self.log_tree.tag_configure('기존', background='#e3f2fd')
        self.log_tree.tag_configure('없음', background='#fff3e0')
        self.log_tree.tag_configure('오류', background='#ffebee')
        # 버튼
        btns = ttk.Frame(parent); btns.pack(fill='x', pady=(5,0))
        for txt, cmd in [("🔄 새로고침", self.refresh_log),
                         ("📤 내보내기", self.export_log),
                         ("🗑️ 초기화",   self.clear_log)]:
            ttk.Button(btns, text=txt, command=cmd).pack(side='left', expand=True, fill='x', padx=2)

    # ─────────────────────────────────────────────────────
    # [D-1] 핫키
    # ─────────────────────────────────────────────────────
    def _setup_hotkeys(self):
        hk_map = {
            'f1':  self.toggle_bot,
            'f2':  self.stop_bot,
            'f3':  lambda: self.capture_coord_hotkey('id_add_button'),
            'f4':  lambda: self.capture_coord_hotkey('chat_button'),
            'f5':  lambda: self.capture_coord_hotkey('profile_area'),
            'f6':  lambda: self.capture_coord_hotkey('confirm_button'),
            'f7':  lambda: self.capture_coord_hotkey('close_button'),
            'f8':  lambda: self.capture_ocr_coord('top_left'),
            'f9':  lambda: self.capture_ocr_coord('bottom_right'),
            'f10': self.ocr_test_capture_and_save,
        }
        for key, func in hk_map.items():
            try: keyboard.add_hotkey(key, func)
            except Exception as e: print(f"[핫키 등록 실패] {key}: {e}")

    # ─────────────────────────────────────────────────────
    # [D-2] 좌표 캡처
    # ─────────────────────────────────────────────────────
    def capture_coord_hotkey(self, key: str):
        try:
            x, y = pyautogui.position()
            self.config['coords'][key] = (x, y)
            self.root.after(0, lambda: self.coord_labels[key].config(
                text=f"X:{x}, Y:{y}", bg='#e8f5e9', fg='#2e7d32')
                if key in self.coord_labels else None)
            self.log_message(f"[핫키] {key}: ({x}, {y})")
        except Exception as e:
            self.log_message(f"[캡처 오류] {key}: {e}")

    def capture_coord(self, key: str):
        self.log_message(f"[좌표] {key} - 3초 후 캡처합니다...")
        def _cap():
            for i in range(3, 0, -1):
                self.log_message(f"  {i}..."); time.sleep(1)
            try:
                x, y = pyautogui.position()
                self.config['coords'][key] = (x, y)
                self.root.after(0, lambda: self.coord_labels[key].config(
                    text=f"X:{x}, Y:{y}", bg='#e8f5e9', fg='#2e7d32')
                    if key in self.coord_labels else None)
                self.log_message(f"  ✅ {key}: ({x}, {y})")
            except Exception as e:
                self.log_message(f"  ❌ 캡처 실패: {e}")
        threading.Thread(target=_cap, daemon=True).start()

    # ─────────────────────────────────────────────────────
    # [D-3] OCR 영역 캡처
    # ─────────────────────────────────────────────────────
    def capture_ocr_coord(self, corner: str):
        try:
            x, y = pyautogui.position()
            if not isinstance(self.config.get('ocr_area'), dict):
                self.config['ocr_area'] = {}
            self.config['ocr_area'][corner] = (x, y)
            if corner == 'top_left':
                self.root.after(0, lambda: self.ocr_topleft_label.config(
                    text=f"X:{x}, Y:{y}", bg='#e8f5e9', fg='#2e7d32'))
                self.log_message(f"[F8] OCR 좌상단: ({x}, {y})")
            else:
                self.root.after(0, lambda: self.ocr_bottomright_label.config(
                    text=f"X:{x}, Y:{y}", bg='#e8f5e9', fg='#2e7d32'))
                self.log_message(f"[F9] OCR 우하단: ({x}, {y})")
            area = self.config['ocr_area']
            if isinstance(area, dict) and 'top_left' in area and 'bottom_right' in area:
                x1,y1 = area['top_left']; x2,y2 = area['bottom_right']
                x1,x2 = min(x1,x2), max(x1,x2)
                y1,y2 = min(y1,y2), max(y1,y2)
                self.config['ocr_area'] = (x1, y1, x2, y2)
                self.log_message(f"[완료] OCR 영역: ({x1},{y1}) ~ ({x2},{y2})")
        except Exception as e:
            self.log_message(f"[OCR 좌표 오류] {e}")

    def capture_ocr_area(self):
        self.log_message("[OCR 드래그] 3초 후 오버레이 표시...")
        def _show():
            time.sleep(3)
            overlay = tk.Toplevel(self.root)
            overlay.attributes('-alpha', 0.30); overlay.attributes('-fullscreen', True)
            overlay.attributes('-topmost', True); overlay.configure(bg='black')
            canvas = tk.Canvas(overlay, cursor='cross', bg='black', highlightthickness=0)
            canvas.pack(fill='both', expand=True)
            canvas.create_text(overlay.winfo_screenwidth()//2, 40,
                text="드래그하여 OCR 영역 선택  (Esc: 취소)",
                fill='white', font=('맑은 고딕',14,'bold'))
            rect = None; sx = sy = 0
            def on_p(e):
                nonlocal sx, sy, rect
                sx, sy = e.x, e.y
                if rect: canvas.delete(rect)
                rect = canvas.create_rectangle(sx, sy, sx, sy, outline='red', width=3)
            def on_d(e):
                if rect: canvas.coords(rect, sx, sy, e.x, e.y)
            def on_r(e):
                x1,y1 = min(sx,e.x),min(sy,e.y); x2,y2 = max(sx,e.x),max(sy,e.y)
                if (x2-x1)<10 or (y2-y1)<10:
                    self.log_message("[OCR 드래그] 영역이 너무 작습니다.")
                    overlay.destroy(); return
                self.config['ocr_area'] = (x1,y1,x2,y2)
                self.root.after(0, lambda: self.ocr_topleft_label.config(
                    text=f"X:{x1}, Y:{y1}", bg='#e8f5e9', fg='#2e7d32'))
                self.root.after(0, lambda: self.ocr_bottomright_label.config(
                    text=f"X:{x2}, Y:{y2}", bg='#e8f5e9', fg='#2e7d32'))
                self.log_message(f"[드래그] OCR 영역: ({x1},{y1}) ~ ({x2},{y2})")
                overlay.destroy()
            def on_esc(e):
                self.log_message("[OCR 드래그] 취소"); overlay.destroy()
            canvas.bind('<Button-1>', on_p)
            canvas.bind('<B1-Motion>', on_d)
            canvas.bind('<ButtonRelease-1>', on_r)
            overlay.bind('<Escape>', on_esc)
            overlay.focus_force()
        threading.Thread(target=_show, daemon=True).start()

    # ─────────────────────────────────────────────────────
    # [B-1] 설정 저장/불러오기/프리셋
    # ─────────────────────────────────────────────────────
    def _sync_config_from_ui(self):
        try:
            for k, var in self.timing_vars.items():
                self.config['timing'][k] = float(var.get())
        except Exception: pass
        try: self.config['csv_path']    = self.csv_entry.get()
        except Exception: pass
        try: self.config['output_path'] = self.output_entry.get()
        except Exception: pass
        try: self.config['counter_start'] = int(self.counter_var.get())
        except Exception: pass
        try: self.config['retry_count']  = int(self.retry_count_var.get())
        except Exception: pass
        try: self.config['log_to_file']  = bool(self.log_to_file_var.get())
        except Exception: pass
        try: self.config['resume_mode']  = bool(self.resume_mode_var.get())
        except Exception: pass
        try: self.config['auto_save_config'] = bool(self.auto_save_var.get())
        except Exception: pass

    def _apply_config_to_ui(self):
        for key, coord in self.config.get('coords', {}).items():
            if key in self.coord_labels:
                if coord and isinstance(coord, (list,tuple)) and len(coord)==2:
                    self.coord_labels[key].config(
                        text=f"X:{coord[0]}, Y:{coord[1]}", bg='#e8f5e9', fg='#2e7d32')
                else:
                    self.coord_labels[key].config(text="미설정", bg='#ffebee', fg='#c62828')
        ocr = self.config.get('ocr_area')
        if ocr and isinstance(ocr, (list,tuple)) and len(ocr)==4:
            x1,y1,x2,y2 = ocr
            self.ocr_topleft_label.config(    text=f"X:{x1}, Y:{y1}", bg='#e8f5e9', fg='#2e7d32')
            self.ocr_bottomright_label.config(text=f"X:{x2}, Y:{y2}", bg='#e8f5e9', fg='#2e7d32')
        else:
            self.ocr_topleft_label.config(    text="미설정", bg='#ffebee', fg='#c62828')
            self.ocr_bottomright_label.config(text="미설정", bg='#ffebee', fg='#c62828')
        for k, var in self.timing_vars.items():
            var.set(self.config['timing'].get(k, DEFAULT_CONFIG['timing'][k]))
        try:
            self.csv_entry.delete(0, tk.END)
            self.csv_entry.insert(0, self.config.get('csv_path',''))
        except Exception: pass
        try:
            self.output_entry.delete(0, tk.END)
            self.output_entry.insert(0, self.config.get('output_path','run_output.xlsx'))
        except Exception: pass
        try: self.counter_var.set(self.config.get('counter_start',1))
        except Exception: pass
        try: self.retry_count_var.set(self.config.get('retry_count',2))
        except Exception: pass
        try: self.log_to_file_var.set(self.config.get('log_to_file',False))
        except Exception: pass
        try: self.resume_mode_var.set(self.config.get('resume_mode',False))
        except Exception: pass
        try: self.auto_save_var.set(self.config.get('auto_save_config',True))
        except Exception: pass

    def save_config(self):
        self._sync_config_from_ui()
        fn = filedialog.asksaveasfilename(title="설정 저장", initialfile="config.json",
            defaultextension=".json", filetypes=[("JSON","*.json"),("All","*.*")])
        if not fn: return
        with open(fn,'w',encoding='utf-8') as f: json.dump(self.config, f, indent=2, ensure_ascii=False)
        try:
            with open('config.json','w',encoding='utf-8') as f: json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception: pass
        messagebox.showinfo("저장", f"설정 저장 완료!\n{fn}")
        self.log_message(f"[설정] 저장: {fn}")

    def load_config(self, silent=False):
        path = None
        if silent and os.path.exists('config.json'):
            path = 'config.json'
        elif not silent:
            path = filedialog.askopenfilename(title="설정 불러오기",
                filetypes=[("JSON","*.json"),("All","*.*")])
            if not path: return
        if not path or not os.path.exists(path):
            if not silent: messagebox.showwarning("알림","불러올 설정 파일이 없습니다.")
            return
        try:
            with open(path,'r',encoding='utf-8') as f: data = json.load(f)
        except Exception as e:
            messagebox.showerror("오류", f"설정 읽기 실패:\n{e}"); return
        merged = make_config_copy()
        deep_merge_into(merged, data)
        self.config = merged
        self._apply_config_to_ui()
        if not silent: messagebox.showinfo("완료", f"설정 불러오기 완료!\n{path}")
        self.log_message(f"[설정] 불러오기: {path}")

    def auto_save_on_exit(self):
        if self.config.get('auto_save_config', True):
            try:
                self._sync_config_from_ui()
                with open('config.json','w',encoding='utf-8') as f:
                    json.dump(self.config, f, indent=2, ensure_ascii=False)
                self.log_message("[자동저장] config.json 저장 완료")
            except Exception as e:
                self.log_message(f"[자동저장 오류] {e}")

    def save_preset(self):
        name = simpledialog.askstring("프리셋 저장","프리셋 이름:")
        if not name: return
        os.makedirs('presets', exist_ok=True)
        self._sync_config_from_ui()
        fn = os.path.join('presets', f"{name}.json")
        with open(fn,'w',encoding='utf-8') as f: json.dump(self.config, f, indent=2, ensure_ascii=False)
        try:
            with open('config.json','w',encoding='utf-8') as f: json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception: pass
        messagebox.showinfo("저장", f"프리셋 '{name}' 저장 완료!")
        self.log_message(f"[프리셋] 저장: {fn}")

    def load_preset(self):
        if not os.path.exists('presets'):
            messagebox.showwarning("경고","저장된 프리셋이 없습니다!"); return
        presets = [os.path.splitext(f)[0] for f in os.listdir('presets') if f.endswith('.json')]
        if not presets:
            messagebox.showwarning("경고","저장된 프리셋이 없습니다!"); return
        dialog = tk.Toplevel(self.root)
        dialog.title("프리셋 선택"); dialog.geometry("300x420"); dialog.resizable(False,False)
        tk.Label(dialog, text="불러올 프리셋 선택:", font=('맑은 고딕',10)).pack(pady=10)
        lb = tk.Listbox(dialog, font=('맑은 고딕',9)); lb.pack(fill='both', expand=True, padx=10, pady=6)
        for p in sorted(presets): lb.insert(tk.END, p)
        def _del():
            sel = lb.curselection()
            if not sel: return
            name = lb.get(sel[0])
            if messagebox.askyesno("삭제 확인", f"프리셋 '{name}'을 삭제하시겠습니까?"):
                try:
                    os.remove(os.path.join('presets', f"{name}.json"))
                    lb.delete(sel[0])
                    self.log_message(f"[프리셋] 삭제: {name}")
                except Exception as e:
                    messagebox.showerror("오류", f"삭제 실패: {e}")
        def _apply():
            sel = lb.curselection()
            if not sel: return
            name = lb.get(sel[0])
            path = os.path.join('presets', f"{name}.json")
            try:
                with open(path,'r',encoding='utf-8') as f: data = json.load(f)
            except Exception as e:
                messagebox.showerror("오류", f"프리셋 읽기 실패:\n{e}"); return
            merged = make_config_copy()
            deep_merge_into(merged, data)
            self.config = merged; self._apply_config_to_ui()
            try:
                with open('config.json','w',encoding='utf-8') as f:
                    json.dump(self.config, f, indent=2, ensure_ascii=False)
            except Exception: pass
            dialog.destroy()
            messagebox.showinfo("완료", f"프리셋 '{name}' 불러오기 완료!")
            self.log_message(f"[프리셋] 불러오기: {path}")
        bf = tk.Frame(dialog); bf.pack(fill='x', padx=10, pady=8)
        ttk.Button(bf, text="불러오기", command=_apply).pack(side='left', expand=True, fill='x', padx=3)
        ttk.Button(bf, text="삭제",     command=_del  ).pack(side='left', expand=True, fill='x', padx=3)
        ttk.Button(bf, text="닫기", command=dialog.destroy).pack(side='left', expand=True, fill='x', padx=3)

    # ─────────────────────────────────────────────────────
    # [E-1] 실행 제어
    # ─────────────────────────────────────────────────────
    def _check_tesseract_ready(self) -> bool:
        try:
            _ = pytesseract.get_tesseract_version(); return True
        except Exception: pass
        if _maybe_set_portable_tesseract():
            try: _ = pytesseract.get_tesseract_version(); return True
            except Exception: pass
        return False

    def toggle_bot(self):
        if not self.is_running: self.start_bot()
        else: self.pause_bot()

    def start_bot(self):
        missing = [k for k,v in self.config['coords'].items() if not v]
        if missing:
            messagebox.showerror("오류", "다음 좌표를 설정하세요:\n" + "\n".join(f"  · {c}" for c in missing)); return
        if not self.config.get('ocr_area'):
            messagebox.showerror("오류", "OCR 영역을 설정하세요! (F8/F9 또는 드래그)"); return
        csv_path = self.csv_entry.get().strip() if hasattr(self,'csv_entry') else ""
        if not csv_path or not os.path.isfile(csv_path):
            messagebox.showerror("오류", "유효한 CSV 파일을 선택하세요!"); return
        if not self._check_tesseract_ready():
            base = _app_dir()
            messagebox.showerror("Tesseract 오류",
                f"Tesseract OCR 실행 파일을 찾을 수 없습니다.\n\n"
                f"권장 구조:\n  {base}\\tesseract\\tesseract.exe\n"
                f"  {base}\\tesseract\\tessdata\\kor.traineddata\n\n"
                "또는 시스템 PATH에 Tesseract를 설치하세요."); return
        self._sync_config_from_ui()
        self.is_running = True; self.is_paused = False
        self.stats = {'신규':0,'기존':0,'없음':0,'오류':0}
        self._start_time = time.time()
        if self.config.get('resume_mode', False):
            self.current_id_index = self.config.get('last_processed_idx', 0)
            self.log_message(f"[이어하기] {self.current_id_index}번째부터 재개합니다.")
        else:
            self.current_id_index = 0
        self.start_btn.config(state='disabled')
        self.pause_btn.config(state='normal')
        self.stop_btn.config(state='normal')
        try:
            self.log_message(f"[디버그] tesseract_cmd = {getattr(pytesseract.pytesseract,'tesseract_cmd',None)}")
            self.log_message(f"[디버그] TESSDATA_PREFIX = {os.environ.get('TESSDATA_PREFIX')}")
            self.log_message(f"[디버그] cwd = {os.getcwd()}")
        except Exception: pass
        self.log_message("[시작] 봇 실행!")
        threading.Thread(target=self.run_bot, daemon=True).start()

    def pause_bot(self):
        if not self.is_running: return
        self.is_paused = not self.is_paused
        if self.is_paused:
            self.pause_btn.config(text="▶️  재개"); self.log_message("[일시정지]")
        else:
            self.pause_btn.config(text="⏸️  일시정지"); self.log_message("[재개]")

    def stop_bot(self):
        self.is_running = False; self.is_paused = False
        self._reset_control_buttons()
        self.log_message("[중지] 봇이 중지되었습니다.")

    def _reset_control_buttons(self):
        try:
            self.start_btn.config(state='normal')
            self.pause_btn.config(state='disabled', text="⏸️  일시정지")
            self.stop_btn.config(state='disabled')
        except Exception: pass

    # ─────────────────────────────────────────────────────
    # [E-2] 메인 루프
    # ─────────────────────────────────────────────────────
    def run_bot(self):
        results = []
        total   = 0
        try:
            csv_path = self.config['csv_path']
            valid, msg = validate_csv_path(csv_path)
            if not valid: raise ValueError(msg)
            df = read_csv_safely(csv_path)
            if 'id' not in df.columns: df.columns = ['id'] + list(df.columns[1:])
            all_ids = [str(x).strip() for x in df['id'].tolist() if str(x).strip()]
            total   = len(all_ids)
            start_idx = self.current_id_index if self.config.get('resume_mode') else 0
            ids = all_ids[start_idx:]
            counter  = self.config.get('counter_start', 1)
            first_id = True
            loop_start = time.time()
            self.log_message(f"[INFO] 총 {total}개 ID, {'이어하기: '+str(start_idx)+'번째부터' if start_idx else '처음부터'} 시작")
            for rel_idx, kakao_id in enumerate(ids):
                abs_idx = start_idx + rel_idx
                if not self.is_running: break
                while self.is_paused and self.is_running: time.sleep(0.3)
                if not self.is_running: break
                progress = (abs_idx+1)/total*100
                eta_str  = self._calc_eta(loop_start, rel_idx+1, len(ids))
                self.root.after(0, lambda p=progress, ai=abs_idx, t=total, e=eta_str:
                    self.update_progress(p, ai, t, e))
                try:
                    self.root.after(0, lambda kid=kakao_id: self.current_id_label.config(text=kid))
                except Exception: pass
                self.log_message(f"\n[{abs_idx+1}/{total}] ID: {kakao_id}")
                result = self.process_id(kakao_id, counter, first_id)
                row = {'ID':kakao_id,'결과':result['status'],'이름':result['name'],
                       '시각':datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
                results.append(row)
                self.stats[result['status']] += 1
                self.root.after(0, self.update_stats)
                status = result['status']
                self.root.after(0, lambda r=row, s=status:
                    self.log_tree.insert('',0, values=(r['ID'],r['결과'],r['이름'],r['시각']), tags=(s,)))
                if result['status'] == '신규': counter += 1; first_id = True
                else: first_id = False
                self.config['last_processed_idx'] = abs_idx+1
                self.current_id_index = abs_idx+1
                time.sleep(0.1)
            out_path = self.save_results(results)
            self.log_message(f"\n[완료] 총 {total}개 처리 완료!")
            summary = (f"총 {total}개 처리 완료\n\n"
                       f"신규: {self.stats['신규']}\n기존: {self.stats['기존']}\n"
                       f"없음: {self.stats['없음']}\n오류: {self.stats['오류']}")
            if out_path: summary += f"\n\n저장 위치:\n{out_path}"
            self.root.after(0, lambda: messagebox.showinfo("완료", summary))
        except Exception as e:
            self.log_message(f"[오류] run_bot 예외: {e}")
            self.root.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self.is_running = False
            self.root.after(0, self._reset_control_buttons)
            try: self.root.after(0, lambda: self.current_id_label.config(text="-"))
            except Exception: pass

    def _calc_eta(self, start_time: float, done: int, total: int) -> str:
        try:
            elapsed = time.time() - start_time
            if done <= 0: return "-"
            avg = elapsed / done
            remaining = avg * (total - done)
            return str(timedelta(seconds=int(remaining)))
        except Exception: return "-"

    # ─────────────────────────────────────────────────────
    # [E-3] ID 처리
    # ─────────────────────────────────────────────────────
    def process_id(self, kakao_id: str, counter: int, first_id: bool) -> dict:
        try:
            timing = self.config['timing']
            coords = self.config['coords']
            if first_id:
                self.log_message("  → Ctrl+A (입력창 초기화)")
                pyautogui.hotkey('ctrl','a'); time.sleep(timing['after_ctrlA'])
                self.log_message("  → [ID로 추가] 클릭")
                pyautogui.click(coords['id_add_button']); time.sleep(timing['after_click'])
            else:
                self.log_message("  → 입력창 클릭")
                pyautogui.click(coords['id_add_button']); time.sleep(timing['after_click'])
                self.log_message("  → 입력창 내용 초기화")
                pyautogui.hotkey('ctrl','a'); time.sleep(0.1)
                pyautogui.press('backspace');  time.sleep(0.2)
            self.log_message(f"  → ID 입력: {kakao_id}")
            pyperclip.copy(kakao_id)
            pyautogui.hotkey('ctrl','v'); time.sleep(0.25)
            pyautogui.press('enter');     time.sleep(timing['after_input'])
            time.sleep(timing['after_ocr'])
            pyautogui.click(coords['id_add_button']); time.sleep(0.6)
            ocr_area = self.config.get('ocr_area')
            if not ocr_area or len(ocr_area) != 4:
                return {'status':'오류','name':'-'}
            x1,y1,x2,y2 = ocr_area
            x1,x2 = min(x1,x2),max(x1,x2); y1,y2 = min(y1,y2),max(y1,y2)
            raw_img = ImageGrab.grab(bbox=(x1,y1,x2,y2))
            self.log_message("  → OCR 분석 중...")
            text_clean, text_raw, text_proc, proc_img = ocr_text_strong(raw_img)
            self.log_message(f"  → OCR 원본:  {text_raw.strip()[:80]}")
            self.log_message(f"  → OCR 정제:  {text_clean[:80]}")
            state = decide_state_from_ocr(text_clean)
            self.log_message(f"  → 판별 결과: [{state}]")
            if state == '신규':
                self.log_message("  ✅ 신규 → 추가 절차 진행")
                return self.add_new_friend(counter, coords, timing)
            if state == '기존':
                self.log_message("  ✅ 기존 → 스킵"); return {'status':'기존','name':'-'}
            if state in ('채널','없음'):
                self.log_message(f"  ⏭️ {state}"); return {'status':'없음','name':'-'}
            self.log_message("  ⚠️ 상태 불명 → 오류")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            os.makedirs('ocr_debug', exist_ok=True)
            raw_img.save(f'ocr_debug/{ts}_{kakao_id}_raw.png')
            proc_img.save(f'ocr_debug/{ts}_{kakao_id}_proc.png')
            return {'status':'오류','name':'-'}
        except pyautogui.FailSafeException:
            self.log_message("  ❌ Fail-safe 발동!"); return {'status':'오류','name':'-'}
        except Exception as e:
            self.log_message(f"  ❌ process_id 예외: {e}"); return {'status':'오류','name':'-'}

    # ─────────────────────────────────────────────────────
    # [E-4] 신규 친구 추가 (재시도 포함)
    # ─────────────────────────────────────────────────────
    def add_new_friend(self, counter: int, coords: dict, timing: dict) -> dict:
        max_retries = self.config.get('retry_count', 2)
        for attempt in range(max_retries + 1):
            if attempt > 0:
                self.log_message(f"  🔄 재시도 {attempt}/{max_retries}...")
                time.sleep(1.0)
            result = self._add_new_friend_core(counter, coords, timing)
            if result['status'] == '신규': return result
            if attempt < max_retries:
                self.log_message(f"  ⚠️ 친구추가 실패 → 재시도 예정")
        self.log_message(f"  ❌ {max_retries+1}회 시도 모두 실패 → 오류")
        return {'status':'오류','name':'-'}

    def _add_new_friend_core(self, counter: int, coords: dict, timing: dict) -> dict:
        try:
            name = f"가망{counter}"
            self.log_message("  → [친구추가] Tab×2 + Enter")
            pyautogui.press('tab'); time.sleep(0.12)
            pyautogui.press('tab'); time.sleep(0.12)
            pyautogui.press('enter'); time.sleep(timing['after_click'])
            self.log_message("  → [1:1 채팅] 클릭")
            pyautogui.click(coords['chat_button']); time.sleep(timing['after_click'])
            self.log_message("  → 프로필 영역 클릭")
            pyautogui.click(coords['profile_area']); time.sleep(timing['after_click'])
            self.log_message("  → 이름 편집 Tab×6 + Enter")
            for _ in range(6):
                pyautogui.press('tab'); time.sleep(timing['after_tab'])
            pyautogui.press('enter'); time.sleep(0.3)
            self.log_message(f"  → 이름 입력: {name}")
            pyautogui.hotkey('ctrl','a'); time.sleep(0.1)
            pyperclip.copy(name); pyautogui.hotkey('ctrl','v'); time.sleep(0.2)
            self.log_message("  → [확인] 클릭")
            pyautogui.click(coords['confirm_button']); time.sleep(timing['after_click'])
            self.log_message("  → [닫기] 클릭")
            pyautogui.click(coords['close_button']); time.sleep(0.4)
            self.log_message(f"  ✅ 신규 친구 저장 완료: {name}")
            return {'status':'신규','name':name}
        except pyautogui.FailSafeException:
            return {'status':'오류','name':'-','_reason':'fail-safe'}
        except Exception as e:
            return {'status':'오류','name':'-','_reason':str(e)}

    # ─────────────────────────────────────────────────────
    # [B-3] Excel 결과 저장
    # ─────────────────────────────────────────────────────
    def save_results(self, results: list):
        try:
            os.makedirs('결과', exist_ok=True)
            ts  = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            out = f"결과/{ts}.xlsx"
            df  = pd.DataFrame(results)
            with pd.ExcelWriter(out, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='결과')
                ws = writer.sheets['결과']
                hf = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                hfont = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
                cm = {
                    '신규': PatternFill(start_color="C8E6C9", fill_type="solid"),
                    '기존': PatternFill(start_color="BBDEFB", fill_type="solid"),
                    '없음': PatternFill(start_color="FFE0B2", fill_type="solid"),
                    '오류': PatternFill(start_color="FFCDD2", fill_type="solid"),
                }
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    s = row[1].value
                    if s in cm:
                        for c in row: c.fill = cm[s]
                for col in ws.columns:
                    ml = max((len(str(cell.value or "")) for cell in col), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(ml+4, 40)
                # 요약 시트 (v3.0)
                summary = {
                    '항목': ['총 처리','신규','기존','없음','오류','저장 시각'],
                    '값':   [len(results),
                             sum(1 for r in results if r.get('결과')=='신규'),
                             sum(1 for r in results if r.get('결과')=='기존'),
                             sum(1 for r in results if r.get('결과')=='없음'),
                             sum(1 for r in results if r.get('결과')=='오류'),
                             ts]
                }
                pd.DataFrame(summary).to_excel(writer, index=False, sheet_name='요약')
                ws2 = writer.sheets['요약']
                for cell in ws2[1]:
                    cell.fill = hf; cell.font = hfont; cell.alignment = Alignment(horizontal='center')
                ws2.column_dimensions['A'].width = 14; ws2.column_dimensions['B'].width = 20
            self.log_message(f"[저장] Excel → {out}")
            if self.config.get('log_to_file', False):
                self._save_log_to_file(results, ts)
            return out
        except Exception as e:
            self.log_message(f"[저장 오류] {e}"); return None

    def _save_log_to_file(self, results: list, ts: str):
        try:
            log_dir = self.config.get('log_dir','logs')
            os.makedirs(log_dir, exist_ok=True)
            lp = os.path.join(log_dir, f"{ts}_log.txt")
            with open(lp,'w',encoding='utf-8') as f:
                f.write(f"카카오톡 친구추가 봇 v3.0 실행 로그\n저장 시각: {ts}\n" + "="*50 + "\n\n")
                for r in results:
                    f.write(f"[{r.get('시각','')}] ID={r.get('ID','')} 결과={r.get('결과','')} 이름={r.get('이름','')}\n")
            self.log_message(f"[로그파일] → {lp}")
        except Exception as e:
            self.log_message(f"[로그파일 오류] {e}")

    # ─────────────────────────────────────────────────────
    # [F-1] 로그 메시지
    # ─────────────────────────────────────────────────────
    def log_message(self, message: str, tag: str = ''):
        if not tag: tag = self._auto_tag(message)
        def _insert():
            if not hasattr(self,'realtime_log') or self.realtime_log is None:
                print(message); return
            try:
                self.realtime_log.config(state='normal')
                if tag: self.realtime_log.insert('end', message+'\n', tag)
                else:   self.realtime_log.insert('end', message+'\n')
                self.realtime_log.see('end')
                self.realtime_log.config(state='disabled')
            except Exception: print(message)
        if threading.current_thread() is threading.main_thread(): _insert()
        else:
            try: self.root.after(0, _insert)
            except Exception: print(message)

    def _auto_tag(self, message: str) -> str:
        m = message.lower()
        if any(k in m for k in ['✅','[완료]','[시작]','신규 친구 저장','저장 완료']): return 'ok'
        if any(k in m for k in ['❌','[오류]','오류','error','exception','fail']): return 'error'
        if any(k in m for k in ['⚠️','[경고]','경고','warn','없음','재시도']): return 'warn'
        if any(k in m for k in ['[디버그]','[info]','→','확인','클릭','ocr']): return 'info'
        return ''

    # ─────────────────────────────────────────────────────
    # [F-2] UI 업데이트
    # ─────────────────────────────────────────────────────
    def update_progress(self, progress: float, current: int, total: int, eta: str = "-"):
        try: self.progress_bar['value'] = progress
        except Exception: pass
        try: self.progress_label.config(text=f"{current+1} / {total}  ({progress:.1f}%)")
        except Exception: pass
        try: self.eta_label.config(text=f"남은 시간: {eta}")
        except Exception: pass

    def update_stats(self):
        try:
            for k, lbl in self.stat_labels.items():
                lbl.config(text=str(self.stats.get(k, 0)))
        except Exception: pass

    def sanitize_paths(self, log: bool = False):
        try:
            cp = self.csv_path_var.get().strip()
            if cp and (not os.path.isfile(cp) or not cp.lower().endswith('.csv')):
                self.csv_path_var.set("")
        except Exception: pass
        try:
            od = self.output_dir_var.get().strip()
            if od and not os.path.isdir(od): self.output_dir_var.set("")
        except Exception: pass
        try:
            xp = self.xlsx_save_var.get().strip()
            if xp:
                folder = os.path.dirname(xp) or ""
                if folder and not os.path.isdir(folder): self.xlsx_save_var.set("")
        except Exception: pass
        if log: self.log_message("[경로] 경로 유효성 정리 완료")

    # ─────────────────────────────────────────────────────
    # [F-3] OCR 테스트
    # ─────────────────────────────────────────────────────
    def ocr_test_capture_and_save(self):
        try:
            ocr_area = self.config.get('ocr_area')
            if not ocr_area or len(ocr_area) != 4:
                messagebox.showerror("OCR 테스트","OCR 영역이 설정되지 않았습니다.\n(F8/F9 또는 드래그)"); return
            x1,y1,x2,y2 = ocr_area
            save_dir = "ocr_captures"; os.makedirs(save_dir, exist_ok=True)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.log_message(f"[OCR 테스트] 영역 캡처: ({x1},{y1}) ~ ({x2},{y2})")
            raw_img  = ImageGrab.grab(bbox=(x1,y1,x2,y2))
            proc_img = preprocess_for_ocr(raw_img)
            raw_path  = os.path.join(save_dir, f"{ts}_raw.png")
            proc_path = os.path.join(save_dir, f"{ts}_proc.png")
            txt_path  = os.path.join(save_dir, f"{ts}.txt")
            raw_img.save(raw_path); proc_img.save(proc_path)
            self.log_message("[OCR 테스트] Tesseract 호출 (kor+eng)...")
            text_clean, text_raw, text_proc, _ = ocr_text_strong(raw_img)
            state = decide_state_from_ocr(text_clean)
            with open(txt_path,"w",encoding="utf-8") as f:
                f.write("=== OCR RAW ===\n"  + (text_raw  or "").strip() + "\n\n")
                f.write("=== OCR PROC ===\n" + (text_proc or "").strip() + "\n\n")
                f.write("=== CLEAN ===\n"    + text_clean + "\n\n")
                f.write("=== 판별 결과 ===\n" + state + "\n")
            self.log_message(f"[OCR 테스트] 저장 완료\n  RAW: {raw_path}\n  PROC: {proc_path}\n  TXT: {txt_path}")
            se = {'신규':'✅','기존':'ℹ️','없음':'⏭️','채널':'📢','오류':'⚠️'}
            messagebox.showinfo("OCR 테스트 결과",
                f"📸 OCR 테스트 결과\n\n"
                f"저장 경로:\n  RAW:  {raw_path}\n  PROC: {proc_path}\n  TXT:  {txt_path}\n\n"
                f"━━━━━━━━━━━━━━━━━━━━\n"
                f"【OCR 원문】\n{(text_raw or '').strip()[:200] or '(비어있음)'}\n\n"
                f"【정제 결과】\n{text_clean[:200] or '(비어있음)'}\n\n"
                f"【판별 결과】 {se.get(state,'')} {state}")
        except Exception as e:
            self.log_message(f"[OCR 테스트 오류] {e}")
            messagebox.showerror("OCR 테스트 오류", str(e))

    # ─────────────────────────────────────────────────────
    # [F-4] 다이얼로그 / 로그 조작
    # ─────────────────────────────────────────────────────
    def show_hotkeys(self):
        messagebox.showinfo("핫키 안내", """\
[실행 제어]
  F1 : 시작 / 재개 / 일시정지 토글
  F2 : 강제 중지

[좌표 설정]
  F3 : ID로 추가 버튼
  F4 : 1:1 채팅 버튼
  F5 : 프로필 영역
  F6 : 확인 버튼
  F7 : X 닫기 버튼

[OCR 영역]
  F8  : 좌상단 좌표 지정
  F9  : 우하단 좌표 지정

[테스트]
  F10 : OCR 테스트 캡처 → 저장 → 결과 확인
""")

    def show_manual(self):
        base  = _app_dir()
        tpath = getattr(pytesseract.pytesseract, 'tesseract_cmd', '(자동 탐지)')
        messagebox.showinfo("사용 설명서 v3.0", f"""\
【기본 사용 순서】
 1) F3~F7  → 각 버튼 좌표 설정
 2) F8/F9 또는 드래그 → OCR 영역 설정
 3) 타이밍 수치 조정
 4) 입력 CSV 파일 선택
 5) 출력 폴더 / 가망 번호 시작값 설정
 6) 설정 저장
 7) F1 또는 [시작] 버튼 클릭

【v3.0 신규 기능】
 · 이어하기 모드 · 재시도 횟수 · 로그 파일 저장
 · ETA 표시 · 요약 시트 · 다크 로그 · 프리셋 삭제

【Tesseract 포터블 구조 (권장)】
 {base}\\tesseract\\tesseract.exe
 {base}\\tesseract\\tessdata\\kor.traineddata

현재 Tesseract 경로: {tpath}
""")

    def refresh_log(self):
        out = self.config.get('output_path','')
        if not out or not os.path.isfile(out): return
        try: df = pd.read_excel(out)
        except Exception: return
        self.log_tree.delete(*self.log_tree.get_children())
        for _, row in df.iterrows():
            s = str(row.get('결과',''))
            self.log_tree.insert('','end',
                values=(row.get('ID',''), row.get('결과',''),
                        row.get('이름',''), row.get('시각', row.get('시간',''))),
                tags=(s,))
        self.log_message(f"[새로고침] {len(df)}건 로드")

    def export_log(self):
        fn = filedialog.asksaveasfilename(title="로그 내보내기",
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx"),("All","*.*")])
        if not fn: return
        data = [self.log_tree.item(i)['values'] for i in self.log_tree.get_children()]
        if not data: messagebox.showwarning("알림","내보낼 데이터가 없습니다."); return
        pd.DataFrame(data, columns=['ID','결과','이름','시각']).to_excel(fn, index=False)
        messagebox.showinfo("완료", f"내보내기 완료:\n{fn}")
        self.log_message(f"[내보내기] {fn}")

    def clear_log(self):
        if not messagebox.askyesno("확인","결과 로그와 통계를 모두 초기화하시겠습니까?"): return
        self.log_tree.delete(*self.log_tree.get_children())
        self.stats = {'신규':0,'기존':0,'없음':0,'오류':0}
        self.update_stats()
        self.log_message("[초기화] 결과 로그 초기화 완료")


# ============================================================
# [G] 메인 진입점
# ============================================================

def main():
    root = tk.Tk()
    app  = KakaoFriendBotPro(root)
    try:
        root.mainloop()
    finally:
        try: keyboard.unhook_all()
        except Exception: pass

if __name__ == "__main__":
    main()
